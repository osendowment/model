"""Pipeline health check — internal consistency of the value/risk pipeline.

Verifies the *derived* CSVs are in sync with their builders and inputs. It
catches the class of staleness bug where an upstream file is regenerated
but a downstream file is not re-run (e.g. complexity.csv changing without
build_workload being re-run — a real bug found mid-development).

Complements `scripts/data_anomalies.py`, which checks metric *values* for
outliers/types; this script checks *consistency*.

Checks:
  1. Each dimension CSV — the four risk dimensions (complexity /
     concentration / security / workload) plus the eligibility-stage builds
     (funding / licenses / active) — matches its `build_<dim>.build()`.
  2. data/risk/risk.csv matches `aggregate_risk.aggregate()` and
     data/eligibility/eligibility.csv matches `build_eligibility.build()`.
  3. data/value/value.csv matches `unify_value_data` (class assignments).
  4. Score-forming component columns are 100%-populated across the risk-scope
     set (a blank = a failed upstream fetch, since edge cases are imputed).
  5. Completeness rule: a component score / the overall risk score is present
     only if ALL of its scored inputs are present (no partial scores on disk).
  6. Long-format git files have no duplicate (repo, sha, metric) keys.
  7. Every GitLab host in the data has a repo_id nickname registered in
     `HOST_NICKNAMES` (src/sources/gitlab/gitlab_client.py).
  8. Every repo_id on disk matches the one unified id shape — `gh/{digits}`,
     `gl/{digits}` (gitlab.com), or `gl/{nickname}-{digits}`.
  9. Id-agreement: every in-scope slug in an id-joined source has rows under
     the repo_id value.csv assigns it (catches silent join-drops when a
     source stamped a different id, e.g. a GitHub mirror of a GitLab repo).
 10. Fetch-date cells in the contract files parse as ISO dates and are not
     future-dated.
 11. data/value/validation.csv and data/value/stats.csv match fresh
     (no-network) builder runs — stats.csv especially, since stats.py
     --check reads it as an input and cannot see it go stale.
 12. The preview deliverables (repos.csv / data.csv / preview.xlsx) match
     their builders; the xlsx is compared by cell content, not bytes.
 13. Curated override rows (value, eligibility, maintainer, cve-package)
     still point at in-scope keys — an orphaned override is a silent no-op.
 14. The npm downloads.status sidecar is well-formed and agrees with
     downloads.csv.
 15. value.csv enum domains: platform in its declared set, git_valid
     strictly True/False.

Usage:
    uv run python scripts/pipeline_health.py
    uv run python scripts/pipeline_health.py --strict   # exit 1 if unhealthy
"""
from __future__ import annotations

import argparse
import csv
import glob
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from rich.console import Console
from rich.table import Table

console = Console()
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# (label, ok, detail)
Result = tuple[str, bool, str]


def _norm(rows: list[dict]) -> dict[str, dict[str, str]]:
    """list[dict] → {repo: {col: str-value}}, None rendered as ''."""
    return {
        r["repo"]: {k: ("" if v is None else str(v)) for k, v in r.items()}
        for r in rows
    }


def _read_csv_by_repo(path: Path) -> dict[str, dict[str, str]]:
    with open(path, encoding="utf-8") as f:
        return {r["repo"]: dict(r) for r in csv.DictReader(f)}


def _set_diff_detail(built: dict, disk: dict) -> str:
    """'repo set differs' detail with the actual members, not just counts."""
    only_built = sorted(set(built) - set(disk))
    only_disk = sorted(set(disk) - set(built))
    bits = [f"builder {len(built)}, disk {len(disk)}"]
    if only_built:
        bits.append("builder-only: " + ", ".join(only_built[:3])
                    + (" …" if len(only_built) > 3 else ""))
    if only_disk:
        bits.append("disk-only: " + ", ".join(only_disk[:3])
                    + (" …" if len(only_disk) > 3 else ""))
    return f"repo set differs ({'; '.join(bits)})"


def _cell_diffs(built: dict, disk: dict) -> int:
    return sum(
        1
        for repo in built
        for k in set(built[repo]) | set(disk[repo])
        if built[repo].get(k, "") != disk[repo].get(k, "")
    )


def check_dimension_csvs() -> list[Result]:
    """Each dimension CSV must equal its builder's current output.

    Covers the four risk dimensions plus the eligibility-stage builds
    (funding / licenses / active, under data/eligibility/).
    """
    from src.eligibility import build_active, build_funding, build_licenses
    from src.risk import (
        build_complexity,
        build_concentration,
        build_security,
        build_workload,
    )

    builders = [
        ("complexity", build_complexity, ROOT / "data" / "risk" / "complexity.csv"),
        ("concentration", build_concentration, ROOT / "data" / "risk" / "concentration.csv"),
        ("security", build_security, ROOT / "data" / "risk" / "security.csv"),
        ("workload", build_workload, ROOT / "data" / "risk" / "workload.csv"),
        ("funding", build_funding, ROOT / "data" / "eligibility" / "funding.csv"),
        ("licenses", build_licenses, ROOT / "data" / "eligibility" / "licenses.csv"),
        ("active", build_active, ROOT / "data" / "eligibility" / "active.csv"),
    ]
    out: list[Result] = []
    for name, mod, path in builders:
        built = _norm(mod.build())
        disk = _read_csv_by_repo(path)
        if set(built) != set(disk):
            out.append((f"{name}.csv", False, _set_diff_detail(built, disk)))
            continue
        diffs = _cell_diffs(built, disk)
        out.append((f"{name}.csv", diffs == 0,
                    "in sync" if diffs == 0
                    else f"{diffs} stale cells — re-run build_{name}"))
    return out


def check_risk_data() -> list[Result]:
    """risk.csv must equal aggregate_risk's join of the dimension CSVs."""
    from src.risk.aggregate_risk import aggregate

    rows = aggregate()
    built = _norm(rows)
    disk = _read_csv_by_repo(ROOT / "data" / "risk" / "risk.csv")
    if set(built) != set(disk):
        return [("risk.csv", False, _set_diff_detail(built, disk))]
    diffs = _cell_diffs(built, disk)
    return [("risk.csv", diffs == 0,
             "in sync" if diffs == 0
             else f"{diffs} stale cells — re-run aggregate_risk")]


def check_eligibility_data() -> list[Result]:
    """eligibility.csv must equal build_eligibility's join of the stage CSVs."""
    from src.eligibility.build_eligibility import build

    built = _norm(build())
    disk = _read_csv_by_repo(ROOT / "data" / "eligibility" / "eligibility.csv")
    if set(built) != set(disk):
        return [("eligibility.csv", False, _set_diff_detail(built, disk))]
    diffs = _cell_diffs(built, disk)
    return [("eligibility.csv", diffs == 0,
             "in sync" if diffs == 0
             else f"{diffs} stale cells — re-run build_eligibility")]


_VALUE_ROWS_CACHE: list[dict] | None = None


def _collect_value_rows() -> list[dict]:
    """unify's per-package input rows, collected once per run (expensive —
    shared by check_value_data and check_overrides_integrity)."""
    global _VALUE_ROWS_CACHE
    if _VALUE_ROWS_CACHE is None:
        from src.value.unify_value_data import ECOSYSTEMS, collect_ecosystem
        rows: list[dict] = []
        for eco in ECOSYSTEMS:
            eco_rows, _ = collect_ecosystem(eco)
            rows.extend(eco_rows)
        _VALUE_ROWS_CACHE = rows
    return _VALUE_ROWS_CACHE


def check_value_data() -> list[Result]:
    """value.csv class assignments must match a fresh unify run."""
    from src.value.unify_value_data import aggregate_by_repo

    built = aggregate_by_repo(_collect_value_rows())
    disk = list(csv.DictReader(open(ROOT / "data" / "value" / "value.csv", encoding="utf-8")))

    if len(built) != len(disk):
        return [("value.csv", False,
                 f"row count differs (unify {len(built)}, disk {len(disk)})")]

    def cls_by_repo(rows: list[dict]) -> dict[str, str]:
        seen = Counter((r.get("repo") or "").strip().lower() for r in rows)
        return {
            (r.get("repo") or "").strip().lower(): r.get("class", "")
            for r in rows
            if (r.get("repo") or "").strip()
            and seen[(r.get("repo") or "").strip().lower()] == 1
        }

    b, d = cls_by_repo(built), cls_by_repo(disk)
    mismatch = sum(1 for g in set(b) & set(d) if b[g] != d[g])
    return [("value.csv", mismatch == 0,
             f"class assignments in sync ({len(set(b) & set(d)):,} repos)"
             if mismatch == 0 else f"{mismatch} class mismatches — re-run unify")]


# Columns that MUST be populated for EVERY risk-scope repo because they form a
# component's risk score. A blank here is a real coverage gap (a failed upstream
# fetch), not a modelling choice — edge cases (dormant / bot-only / new repos)
# are imputed by the builder, so the only way these go blank is missing data.
# This is the guard that catches "late joiner" gaps: a fetcher auto-scopes to
# load_top_repos() but smart-skips cached repos, so a repo that joins class-A
# scope after the last fetch silently misses its scores until a forced re-fetch.
# Extend per dimension as 100%-coverage guarantees are added.
# workload is included: a zero-active-contributor repo is scored with AC=1
# (flagged `dormant`) rather than abstaining, so every top repo gets a score.
SCORE_COMPONENT_COVERAGE: dict[str, list[str]] = {
    "risk/concentration.csv": ["bf_commits_git_5y", "hhi_commits_git_5y", "score"],
    "risk/complexity.csv": ["score"],
    "risk/security.csv": ["score"],
    "eligibility/funding.csv": ["score"],
    "risk/workload.csv": ["score"],
}

def check_score_component_coverage() -> list[Result]:
    """Score-forming component columns must be 100% across the risk-scope set.

    For concentration, the builder imputes dormant / bot-only / new repos
    (bus factor 1, HHI 10000), so a blank `bf_commits_git_5y` / `hhi_commits_git_5y`
    / `score` means an upstream git fetch failed — surfaced here as the gap to
    fix (raise the fetch `--timeout` or re-run the fetcher for the listed repos).
    Empty-tree repos (scc measured 0 loc at every snapshot — archived stubs
    stripped to a README) no longer need an exemption: build_complexity scores
    them as measured zeros, so complexity/workload must cover every repo too.
    """
    from src.common.repos import load_top_repos

    repos = [e.repo for e in load_top_repos()]
    out: list[Result] = []
    for fname, cols in SCORE_COMPONENT_COVERAGE.items():
        disk = _read_csv_by_repo(ROOT / "data" / fname)
        scope = repos
        n = len(scope)
        note = ""
        for col in cols:
            missing = [r for r in scope
                       if not str(disk.get(r, {}).get(col, "")).strip()]
            present = n - len(missing)
            ok = not missing
            if ok:
                detail = f"{present}/{n} (100%{note})"
            else:
                shown = ", ".join(missing[:5]) + (" …" if len(missing) > 5 else "")
                detail = f"{present}/{n} — {len(missing)} missing: {shown}{note}"
            out.append((f"{fname}:{col}", ok, detail))
    return out


# A score may be present ONLY if all of its scored inputs are present. For each
# component CSV the inputs are the `composite_cols` its builder geometric-means
# (src/risk/build_<dim>.py); for risk.csv they are the four component scores
# aggregate_risk geometric-means. The producers already enforce this (a missing
# input blanks the score), so any score-with-a-missing-input on disk is a real
# inconsistency — a hand-edit or a builder/aggregator regression.
SCORE_INPUTS: dict[str, list[str]] = {
    "risk/concentration.csv": ["bf_commits_git_5y_p", "hhi_commits_git_5y_p"],
    "risk/complexity.csv":    ["loc_eoy_p", "cyclomatic_max_p"],
    "risk/security.csv":      ["openssf_score_p", "cve_score"],
    "eligibility/funding.csv": ["gh_sponsorships_p", "oc_avg_funding_p"],
    "risk/workload.csv":      ["loc_per_ac_p", "cve_per_ac_p", "nni_per_ac_p"],
    "risk/risk.csv":          ["concentration", "complexity", "security", "workload"],
}

# Files whose builder composes with `max_composite_any` (worst-of over whichever
# axes are present) instead of the strict all-inputs geometric mean. For these a
# score needs AT LEAST ONE input, not all of them: security scores GitLab repos
# without a Scorecard from the CVE axis alone (see src/risk/build_security.py).
SCORE_INPUTS_ANY: set[str] = {"risk/security.csv"}

# Column holding each file's aggregated score (default "score"; the top-level
# risk.csv names it "risk_score", value.csv would be "value_score").
SCORE_COLUMN: dict[str, str] = {"risk/risk.csv": "risk_score"}


def check_score_input_completeness() -> list[Result]:
    """A component score / the risk score may be present only if ALL its scored
    inputs are present — except `SCORE_INPUTS_ANY` files, whose worst-of
    (`max_composite_any`) composite needs at least ONE input.

    Enforces the completeness rule end to end: each component score is the
    geometric mean of its `composite_cols` (blanked when any is missing), and the
    overall risk score is the geometric mean of the four component scores (blanked
    when any component is missing). A row carrying a score while its required
    inputs are blank violates the rule.
    """
    out: list[Result] = []
    for fname, inputs in SCORE_INPUTS.items():
        path = ROOT / "data" / fname
        if not path.exists():
            out.append((f"{fname} score⟸inputs", False, "file missing"))
            continue
        rows = list(csv.DictReader(open(path, encoding="utf-8")))
        score_col = SCORE_COLUMN.get(fname, "score")
        scored = [r for r in rows if str(r.get(score_col, "")).strip()]
        need = any if fname in SCORE_INPUTS_ANY else all
        bad = [r.get("repo", "?") for r in scored
               if not need(str(r.get(c, "")).strip() for c in inputs)]
        ok = not bad
        if ok:
            rule = "≥1 input present (worst-of)" if fname in SCORE_INPUTS_ANY \
                else "every input present"
            detail = f"{len(scored):,} scored rows — {rule}"
        else:
            shown = ", ".join(bad[:5]) + (" …" if len(bad) > 5 else "")
            detail = f"{len(bad)} score(s) with a missing input: {shown}"
        out.append((f"{fname} score⟸inputs", ok, detail))
    return out


def check_long_format_keys() -> list[Result]:
    """Long-format git files must have unique (repo, sha, metric) keys."""
    out: list[Result] = []
    for path in sorted(glob.glob(str(ROOT / "data" / "sources" / "git" / "*.csv"))):
        rows = list(csv.DictReader(open(path, encoding="utf-8")))
        if not rows or not {"repo", "commit_sha", "metric"} <= set(rows[0]):
            continue  # not a long-format file
        keys = Counter((r["repo"], r["commit_sha"], r["metric"]) for r in rows)
        dups = sum(1 for v in keys.values() if v > 1)
        name = Path(path).name
        out.append((f"git/{name}", dups == 0,
                    f"{len(rows):,} rows, keys unique" if dups == 0
                    else f"{dups} duplicate (repo,sha,metric) keys"))
    return out


def check_value_criticality() -> list[Result]:
    """Every valid class-A GitHub row in value.csv carries an openssf_crit score.

    The `openssf_crit` column is filled by `src.value.apply_criticality` from
    the OpenSSF criticality fetch, whose scope is exactly this set (archived
    included) — so a blank is a missing/failed fetch or a skipped apply step,
    never "not applicable".
    """
    disk = list(csv.DictReader(open(ROOT / "data" / "value" / "value.csv",
                                    encoding="utf-8")))
    gate = [r for r in disk
            if (r.get("platform") or "").lower() == "github"
            and (r.get("git_valid") or "") == "True"
            and (r.get("class") or "") == "A"]
    blank = [r["repo"] for r in gate if not (r.get("openssf_crit") or "").strip()]
    if blank:
        sample = ", ".join(blank[:3]) + ("…" if len(blank) > 3 else "")
        return [("value.csv:openssf_crit", False,
                 f"{len(gate) - len(blank)}/{len(gate)} — {len(blank)} blank: {sample}")]
    return [("value.csv:openssf_crit", True,
             f"{len(gate)}/{len(gate)} valid class-A github rows scored")]


# Source CSVs whose rows the builders join by the stable repo_id. A fetcher
# rewrite that drops or blanks the column silently blanks entire dimensions
# (this happened: contributor-commits, commits-years, and the GitHub
# contributor file were all clobbered by schema-unaware rewrites in one run).
ID_JOINED_SOURCES = [
    "sources/git/contributor-commits.csv",
    "sources/github/contributor-commits.csv",
    "sources/git/scc.csv",
    "sources/git/lizard.csv",
    "sources/git/openssf.csv",
    "sources/git/commits-years.csv",
    "sources/git/depsdev.csv",
    "sources/github/issues.csv",
    "sources/osv/cves.csv",
]


def _value_id_by_slug() -> dict[str, str]:
    """{repo slug lowercased: repo_id} from value.csv (blank ids skipped)."""
    out: dict[str, str] = {}
    with open(ROOT / "data" / "value" / "value.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            slug = (r.get("repo") or "").strip().lower()
            rid = (r.get("repo_id") or "").strip()
            if slug and rid:
                out[slug] = rid
    return out


def check_source_value_id_agreement() -> list[Result]:
    """Every in-scope slug's rows in an id-joined source include the repo_id
    value.csv currently assigns to that slug.

    The builders join these files by the value-stage id, so a slug whose rows
    all carry a *different* id (e.g. the GitHub-mirror id of a repo whose
    canonical home resolved to GitLab) is silently invisible to every join —
    the dimension goes blank with no error. Rows under a stale id are
    tolerated when rows under the current id also exist (dead weight, joins
    fine); the failure is only when the current id is entirely absent.
    Repair: scripts/heal_repo_id_drift.py restamps drifted-only slugs.
    """
    from src.common.repos import load_top_repos

    scope = {e.repo for e in load_top_repos(skip_archived=False)}
    value_id = _value_id_by_slug()
    out: list[Result] = []
    for rel in ID_JOINED_SOURCES:
        path = ROOT / "data" / rel
        if not path.exists():
            out.append((f"{rel}:id agreement", False, "file missing"))
            continue
        ids_by_slug: dict[str, set[str]] = {}
        with open(path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                slug = (row.get("repo") or "").strip().lower()
                rid = (row.get("repo_id") or "").strip()
                if rid and slug in scope and slug in value_id:
                    ids_by_slug.setdefault(slug, set()).add(rid)
        drifted = sorted(s for s, ids in ids_by_slug.items() if value_id[s] not in ids)
        stale_extra = sum(1 for s, ids in ids_by_slug.items()
                          if value_id[s] in ids and len(ids) > 1)
        note = f" ({stale_extra} slug(s) also carry stale extra ids)" if stale_extra else ""
        if drifted:
            shown = ", ".join(drifted[:3]) + (" …" if len(drifted) > 3 else "")
            out.append((f"{rel}:id agreement", False,
                        f"{len(drifted)} in-scope slug(s) with no row under the "
                        f"value.csv id: {shown}{note}"))
        else:
            out.append((f"{rel}:id agreement", True,
                        f"{len(ids_by_slug)} in-scope slugs agree with value.csv{note}"))
    return out


def check_source_repo_id_integrity() -> list[Result]:
    """Every id-joined source CSV keeps its repo_id column populated for
    in-scope repos.

    Two failure modes, both observed in the wild:
      1. header lost the repo_id column entirely (schema-unaware rewrite);
      2. rows for in-scope repos carry a blank repo_id (id map lagged a
         rename), so the id-keyed builder joins silently drop them.
    Out-of-scope rows may legitimately be blank (legacy/orphan slugs), so
    the gate only covers slugs in the current risk scope.
    """
    from src.common.repos import load_top_repos

    scope = {e.repo for e in load_top_repos(skip_archived=False)}
    out: list[Result] = []
    for rel in ID_JOINED_SOURCES:
        path = ROOT / "data" / rel
        if not path.exists():
            out.append((f"{rel}:repo_id", False, "file missing"))
            continue
        with open(path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if "repo_id" not in (reader.fieldnames or []):
                out.append((f"{rel}:repo_id", False,
                            "repo_id column MISSING from header (schema clobbered)"))
                continue
            blank: set[str] = set()
            for row in reader:
                slug = (row.get("repo") or "").strip().lower()
                if slug in scope and not (row.get("repo_id") or "").strip():
                    blank.add(slug)
        if blank:
            sample = ", ".join(sorted(blank)[:3]) + ("…" if len(blank) > 3 else "")
            out.append((f"{rel}:repo_id", False,
                        f"{len(blank)} in-scope repo(s) with blank repo_id: {sample}"))
        else:
            out.append((f"{rel}:repo_id", True, "in-scope rows fully id-keyed"))
    return out


# Repo-keyed source schema contract (see CLAUDE.md ## Auditability): every
# fetched source CSV keyed by a GitHub repo name must carry repo_id AND a
# fetch-date column — per row, or in a documented per-repo .status sidecar.
# Value: the date column name in the file itself, or ("sidecar", <path>, <col>).
SOURCE_SCHEMA_CONTRACT: dict[str, object] = {
    "sources/git/contributor-commits.csv":
        ("sidecar", "sources/git/contributor-commits.status.csv", "fetched_at"),
    "sources/github/contributor-commits.csv":
        ("sidecar", "sources/github/contributor-commits.status.csv", "fetched_at"),
    "sources/osv/cves.csv": ("sidecar", "sources/osv/queried.csv", "fetched_at"),
    "sources/git/commits-years.csv": "fetched_at",
    "sources/github/issues.csv": "fetched_at",
    "sources/git/scc.csv": "checked_at",
    "sources/git/lizard.csv": "checked_at",
    "sources/git/openssf.csv": "checked_at",
    "sources/git/depsdev.csv": "checked_at",
    "sources/openssf/checks.csv": "date",
    "sources/openssf/criticality.csv": "checked_at",
    "sources/github/repos.csv": "fetched_at",
    "sources/github/sponsors.csv": "fetched_at",
    "sources/github/funding-yml.csv": "fetched_at",
    "sources/gitlab/funding-files.csv": "fetched_at",
    "sources/npm/funding.csv": "fetched_at",
    "sources/pypi/funding.csv": "fetched_at",
    "sources/funding/host-by-repo.csv": "host_checked",
    "sources/ossfuzz/projects.csv": "fetched_at",
    "sources/floss-fund/funding-json.csv": "fetched_at",
    "sources/github/canonical-repos.csv": "fetched_at",
}


def _header(path) -> list[str]:
    with open(path, encoding="utf-8") as f:
        return next(csv.reader(f), [])


def check_source_schema_contract() -> list[Result]:
    """Repo-keyed fetched sources carry repo_id + a fetch-date column.

    Header-presence check only (cheap): fullness of repo_id for in-scope rows
    is covered by check_source_repo_id_integrity; date cells may legitimately
    be blank for rows written before date-tracking existed.
    """
    out: list[Result] = []
    for rel, date_spec in SOURCE_SCHEMA_CONTRACT.items():
        path = ROOT / "data" / rel
        if not path.exists():
            out.append((f"{rel}:schema", False, "file missing"))
            continue
        cols = _header(path)
        problems = []
        if "repo_id" not in cols:
            problems.append("no repo_id column")
        if isinstance(date_spec, tuple):
            _, side_rel, side_col = date_spec
            side = ROOT / "data" / side_rel
            if not side.exists():
                problems.append(f"status sidecar missing ({side_rel})")
            elif side_col not in _header(side):
                problems.append(f"sidecar lacks {side_col}")
        elif date_spec not in cols:
            problems.append(f"no {date_spec} column")
        if problems:
            out.append((f"{rel}:schema", False, "; ".join(problems)))
        else:
            out.append((f"{rel}:schema", True, "repo_id + fetch date present"))
    return out


def check_gitlab_host_nicknames() -> list[Result]:
    """Every GitLab host present in the data has a repo_id nickname.

    `make_repo_id` builds ids as `gl/{nickname}-{project_id}` from
    `HOST_NICKNAMES` (src/sources/gitlab/gitlab_client.py) and raises on an
    unmapped host — so a host that reaches the data without a nickname means
    the map must be extended BY HAND (never auto-derive a nickname: an
    invented id would change once the real nickname lands). Hosts are
    collected from every `host` column under data/sources/gitlab/ and from
    every GitLab git_url in value.csv (covers projects that never resolved
    to a repos.csv row). Also guards the map itself: nicknames must be
    unique, lowercase alphanumeric, and non-numeric (a numeric nickname
    would collide with bare gitlab.com ids).
    """
    from src.sources.gitlab.gitlab_client import HOST_NICKNAMES, parse_git_url

    out: list[Result] = []
    nicks = [n for n in HOST_NICKNAMES.values() if n]
    bad_nicks = [n for n in nicks if not re.fullmatch(r"[a-z][a-z0-9]*", n)]
    dup_nicks = [n for n, c in Counter(nicks).items() if c > 1]
    map_problems = []
    if bad_nicks:
        map_problems.append(f"malformed nickname(s): {', '.join(bad_nicks)}")
    if dup_nicks:
        map_problems.append(f"duplicate nickname(s): {', '.join(dup_nicks)}")
    out.append(("gitlab:HOST_NICKNAMES", not map_problems,
                "; ".join(map_problems) if map_problems
                else f"{len(HOST_NICKNAMES)} hosts mapped, nicknames well-formed"))

    hosts: set[str] = set()
    for path in sorted(glob.glob(str(ROOT / "data" / "sources" / "gitlab" / "*.csv"))):
        with open(path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if "host" not in (reader.fieldnames or []):
                continue
            hosts.update(h for r in reader if (h := (r.get("host") or "").strip().lower()))
    with open(ROOT / "data" / "value" / "value.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            parsed = parse_git_url(r.get("git_url") or "")
            if parsed:
                hosts.add(parsed[0])
    missing = sorted(hosts - set(HOST_NICKNAMES))
    out.append(("gitlab:host nicknames", not missing,
                f"{len(missing)} host(s) without a nickname — add manually to "
                f"HOST_NICKNAMES in src/sources/gitlab/gitlab_client.py: "
                f"{', '.join(missing)}" if missing
                else f"{len(hosts)} host(s) in data, all nicknamed"))
    return out


def check_repo_id_shape() -> list[Result]:
    """Every repo_id on disk matches the one unified id shape.

    Allowed forms: blank, `gh/{digits}`, bare `gl/{digits}` (gitlab.com), or
    `gl/{nickname}-{digits}` with a nickname registered in `HOST_NICKNAMES`.
    Everything else fails — bare numerics (a fetcher writing the raw API id),
    host-form gl/ ids, or garbage. Scans the repo_id column of every CSV
    under data/ that has one (header sniff first, so vendor dumps and non-id
    files cost nothing). Repair: scripts/heal_repo_id_drift.py.
    """
    from src.sources.gitlab.gitlab_client import HOST_NICKNAMES

    nicks = sorted(n for n in HOST_NICKNAMES.values() if n)
    valid = re.compile(rf"gh/\d+|gl/(?:(?:{'|'.join(map(re.escape, nicks))})-)?\d+")
    out: list[Result] = []
    scanned = 0
    for path in sorted((ROOT / "data").rglob("*.csv")):
        if "repo_id" not in _header(path):
            continue
        scanned += 1
        bad: set[str] = set()
        with open(path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                rid = (row.get("repo_id") or "").strip()
                if rid and not valid.fullmatch(rid):
                    bad.add(rid)
        if bad:
            sample = ", ".join(sorted(bad)[:3]) + (" …" if len(bad) > 3 else "")
            out.append((f"{path.relative_to(ROOT)}:id-shape", False,
                        f"{len(bad)} non-canonical repo_id(s): {sample}"))
    out.append(("repo_id shape", not out or all(ok for _, ok, _ in out),
                f"all repo_ids canonical across {scanned} repo_id-keyed files"
                if not [1 for _, ok, _ in out if not ok] else
                "non-canonical repo_ids found (see rows above)"))
    return out


def _parse_fetch_date(raw: str):
    """ISO-8601 → aware datetime (naive treated as UTC); None if unparseable."""
    from datetime import timezone
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def check_fetch_date_sanity() -> list[Result]:
    """Every fetch-date cell in the contract files is a parseable ISO date
    that is not in the future.

    The freshness gates (`src/common/freshness.py`) treat any timestamp
    >= cutoff as fresh with no upper bound, so a future-dated or malformed
    cell silently pins a row fresh forever (or stale forever) — this is the
    audit the gates don't do. Blank cells are allowed (rows predating date
    tracking); only non-blank cells are validated. +1 day of tolerance
    absorbs timezone skew.
    """
    from datetime import timedelta, timezone

    horizon = datetime.now(timezone.utc) + timedelta(days=1)
    files: dict[str, str] = {}
    for rel, date_spec in SOURCE_SCHEMA_CONTRACT.items():
        if isinstance(date_spec, tuple):
            _, side_rel, side_col = date_spec
            files[side_rel] = side_col
        else:
            files[rel] = date_spec
    out: list[Result] = []
    bad_files = 0
    for rel, col in sorted(files.items()):
        path = ROOT / "data" / rel
        if not path.exists():
            continue  # missing files are check_source_schema_contract's finding
        malformed: set[str] = set()
        future: set[str] = set()
        with open(path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                raw = (row.get(col) or "").strip()
                if not raw:
                    continue
                dt = _parse_fetch_date(raw)
                if dt is None:
                    malformed.add(raw)
                elif dt > horizon:
                    future.add(raw)
        if malformed or future:
            bad_files += 1
            bits = []
            if malformed:
                bits.append(f"{len(malformed)} malformed (e.g. {sorted(malformed)[0]!r})")
            if future:
                bits.append(f"{len(future)} future-dated (e.g. {sorted(future)[-1]!r})")
            out.append((f"{rel}:{col}", False, "; ".join(bits)))
    out.append(("fetch dates sane", bad_files == 0,
                f"{len(files)} dated files — all cells parseable, none future"
                if bad_files == 0 else f"{bad_files} file(s) with bad date cells"))
    return out


def check_validation_data() -> list[Result]:
    """validation.csv must equal a fresh (no-network) build_validation run.

    Replicates main()'s pure path — collect targets from value.csv, verdicts
    from the source caches, override pins on top — without the ls-remote
    refresh or the write-back. build()'s hard gate (a target with no verdict)
    is reproduced as a FAIL here instead of a SystemExit.
    """
    from src.value.build_validation import (
        VALIDATION_FIELDS,
        VALUE_FILE,
        apply_overrides,
        build,
        collect_targets,
        load_verdicts,
    )
    from src.value.unify_value_data import load_repo_overrides

    with open(VALUE_FILE, encoding="utf-8") as f:
        value_rows = list(csv.DictReader(f))
    targets = collect_targets(value_rows)
    verdicts = apply_overrides(load_verdicts(), load_repo_overrides())
    missing = sorted(k for k in targets if k not in verdicts)
    if missing:
        shown = ", ".join(f"{t} [{ty}]" for t, ty in missing[:3])
        return [("validation.csv", False,
                 f"{len(missing)} target(s) with no verdict: {shown} — "
                 "run the value rollup")]
    built_rows, _ = build(targets, verdicts)
    built = {(r["target"], r["type"]): r for r in built_rows}
    with open(ROOT / "data" / "value" / "validation.csv", encoding="utf-8") as f:
        disk = {(r["target"], r["type"]): dict(r) for r in csv.DictReader(f)}
    if set(built) != set(disk):
        only_b = len(set(built) - set(disk))
        only_d = len(set(disk) - set(built))
        return [("validation.csv", False,
                 f"target set differs (builder {len(built)}, disk {len(disk)}; "
                 f"{only_b} builder-only, {only_d} disk-only) — re-run build_validation")]
    diffs = sum(1 for k in built
                for col in VALIDATION_FIELDS
                if built[k].get(col, "") != disk[k].get(col, ""))
    return [("validation.csv", diffs == 0,
             f"in sync ({len(built):,} targets)" if diffs == 0
             else f"{diffs} stale cells — re-run build_validation")]


def check_stats_data() -> list[Result]:
    """stats.csv must equal a fresh (no-network) build_stats run.

    stats.csv feeds `scripts/stats.py` and `src/common/params.py` as an
    *input*, and `stats.py --check` recomputes docs figures FROM it — so a
    stale stats.csv self-certifies there; this is the only guard that
    re-derives it. npm downloads come from the on-disk year cache instead of
    `load_npm()` (which fetches missing years); if the cache is behind, the
    mismatch surfaces here and a real `build_stats` run repairs both.
    """
    from src.sources.npm.fetch_npm_stats import _load_existing
    from src.value.build_stats import (
        ALL_ECOSYSTEMS,
        DOWNLOAD_ECOSYSTEMS,
        DOWNLOAD_LOADERS,
        build_stats_rows,
        package_counts,
    )

    downloads = {
        eco: (_load_existing() if eco == "npm" else DOWNLOAD_LOADERS[eco]())
        for eco in DOWNLOAD_ECOSYSTEMS
    }
    counts = {eco: package_counts(eco) for eco in ALL_ECOSYSTEMS}
    built = {r["metric"]: {k: str(v) for k, v in r.items()}
             for r in build_stats_rows(downloads, counts)}
    with open(ROOT / "data" / "value" / "stats.csv", encoding="utf-8") as f:
        disk = {r["metric"]: dict(r) for r in csv.DictReader(f)}
    if set(built) != set(disk):
        return [("stats.csv", False, _set_diff_detail(built, disk))]
    diffs = sum(1 for m in built
                for eco in ALL_ECOSYSTEMS
                if built[m].get(eco, "") != disk[m].get(eco, ""))
    return [("stats.csv", diffs == 0,
             f"in sync ({len(built)} metrics)" if diffs == 0
             else f"{diffs} stale cells — re-run build_stats")]


def check_preview_data() -> list[Result]:
    """The preview deliverables are in sync end to end.

    repos.csv vs build_results.build(), data.csv vs build_data.build(), and
    preview.xlsx cell contents vs the CSV-backed sheets. The xlsx is compared by
    *content* (openpyxl), never by bytes — Workbook.save() embeds timestamps, so
    bytes are non-deterministic.
    """
    from openpyxl import load_workbook

    from src.preview.build_preview_workbook import (
        OUTPUT_FILE,
        REPOS_DROP_COLS,
        SHEETS,
        _cell_value,
    )
    from src.preview.build_results import build as build_repos_rows

    out: list[Result] = []

    built = {r["repo_id"]: {k: ("" if v is None else str(v)) for k, v in r.items()}
             for r in build_repos_rows()}
    disk = {}
    with open(ROOT / "data" / "preview" / "repos.csv", encoding="utf-8") as f:
        disk = {r["repo_id"]: dict(r) for r in csv.DictReader(f)}
    if set(built) != set(disk):
        out.append(("preview/repos.csv", False, _set_diff_detail(built, disk)))
    else:
        diffs = _cell_diffs(built, disk)
        out.append(("preview/repos.csv", diffs == 0,
                    f"in sync ({len(built)} repos)" if diffs == 0
                    else f"{diffs} stale cells — re-run build_results"))

    from src.preview.build_data import build as build_data_rows

    built = {r["repo_id"]: {k: ("" if v is None else str(v)) for k, v in r.items()}
             for r in build_data_rows()}
    with open(ROOT / "data" / "preview" / "data.csv", encoding="utf-8") as f:
        disk = {r["repo_id"]: dict(r) for r in csv.DictReader(f)}
    if set(built) != set(disk):
        out.append(("preview/data.csv", False, _set_diff_detail(built, disk)))
    else:
        diffs = _cell_diffs(built, disk)
        out.append(("preview/data.csv", diffs == 0,
                    f"in sync ({len(built)} repos)" if diffs == 0
                    else f"{diffs} stale cells — re-run build_data"))

    if not OUTPUT_FILE.exists():
        out.append(("preview/preview.xlsx", False, "file missing — re-run build_preview_workbook"))
        return out
    wb = load_workbook(OUTPUT_FILE, read_only=True)
    stale = []
    for sheet, csv_path in SHEETS:
        with open(csv_path, encoding="utf-8", newline="") as f:
            expected = [[_cell_value(c) for c in row] for row in csv.reader(f)]
        # The repos sheet drops its join keys (REPOS_DROP_COLS) for humans —
        # strip the same columns from the CSV rows before comparing.
        if sheet == "repos" and expected:
            drop = {i for i, name in enumerate(expected[0]) if name in REPOS_DROP_COLS}
            expected = [[c for i, c in enumerate(row) if i not in drop]
                        for row in expected]
        if sheet not in wb.sheetnames:
            stale.append(f"{sheet}: sheet missing")
            continue
        actual = [list(row) for row in wb[sheet].iter_rows(values_only=True)]
        if len(actual) != len(expected):
            stale.append(f"{sheet}: {len(actual)-1} rows vs {len(expected)-1} in CSV")
        else:
            cells = sum(1 for a, e in zip(actual, expected)
                        if list(a) + [None] * (len(e) - len(a)) != e)
            if cells:
                stale.append(f"{sheet}: {cells} differing row(s)")
    wb.close()
    out.append(("preview/preview.xlsx", not stale,
                f"{len(SHEETS)} CSV sheets match their CSVs" if not stale
                else "; ".join(stale) + " — re-run build_preview_workbook"))
    return out


def check_overrides_integrity() -> list[Result]:
    """Every curated override row still points at something that exists.

    Overrides are silent no-ops when their key no longer matches — a renamed
    slug, a package that left the dep tree, a repo that left scope — so rot
    is invisible until someone notices a correction stopped applying. Orphan
    rules per file (mirroring each consumer's join):
      - value/overrides.csv: (package, ecosystem) must appear in unify's
        per-package input rows.
      - eligibility/overrides.csv: per-repo rows by repo_id in the top scope
        (archived included); `owner/*` rows by owner having an in-scope repo.
      - eligibility/maintainer-overrides.csv: login is a bus-factor
        maintainer of an in-scope repo.
      - risk/cve-package-overrides.csv: github_repo (rename-canonicalized)
        is an in-scope slug.
    """
    from src.common.repos import canonical_repo_map, load_top_repos
    from src.value.unify_value_data import load_repo_overrides

    out: list[Result] = []
    top = load_top_repos(skip_archived=False)
    scope_slugs = {e.repo for e in top}
    scope_ids = {e.repo_id for e in top if e.repo_id}
    scope_owners = {s.split("/", 1)[0] for s in scope_slugs}

    value_pkgs = {(r.get("package", ""), r.get("ecosystem", ""))
                  for r in _collect_value_rows()}
    orphans = sorted(f"{p}/{e}" for (p, e) in load_repo_overrides() if (p, e) not in value_pkgs)
    out.append(("value/overrides.csv", not orphans,
                f"{len(orphans)} orphan key(s): " + ", ".join(orphans[:3])
                + (" …" if len(orphans) > 3 else "") if orphans
                else "every (package, ecosystem) key matches an input row"))

    # Rows whose key left the TOP scope but still matches a repo in the full
    # value table are DORMANT, not orphaned: repos oscillate around the class
    # boundary, and the curation re-applies the moment they re-enter scope.
    # Only a key matching nothing anywhere (typo / unhandled rename) fails.
    all_ids: set[str] = set()
    all_owners: set[str] = set()
    with open(ROOT / "data" / "value" / "value.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if (rid := (r.get("repo_id") or "").strip()):
                all_ids.add(rid)
            all_owners.add((r.get("repo") or "").split("/", 1)[0].lower())

    bad: list[str] = []
    dormant: list[str] = []
    n = 0
    with open(ROOT / "data" / "eligibility" / "overrides.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            n += 1
            slug = (r.get("repo") or "").strip().lower()
            if slug.endswith("/*"):
                owner = slug[:-2]
                if owner in scope_owners:
                    continue
                (dormant if owner in all_owners else bad).append(slug)
            else:
                rid = (r.get("repo_id") or "").strip()
                if rid in scope_ids:
                    continue
                (dormant if rid in all_ids else bad).append(slug)
    detail = (f"{len(bad)} orphan row(s): " + ", ".join(bad[:3])
              + (" …" if len(bad) > 3 else "")) if bad else \
             f"all {n} rows resolve" + \
             (f" ({len(dormant)} dormant, out of top scope: " +
              ", ".join(dormant[:3]) + (" …" if len(dormant) > 3 else "") + ")"
              if dormant else "")
    out.append(("eligibility/overrides.csv", not bad, detail))

    # The override only ever fires through build_funding's bus-factor join, so
    # the membership universe IS load_bf_contributors() — a login absent from it
    # can never flip bf_maintainer_fundable, whatever the curation says.
    from src.sources.github.bf_contributors import load_bf_contributors

    bf_logins = {login.lower() for logins in load_bf_contributors().values()
                 for login in logins}
    with open(ROOT / "data" / "eligibility" / "maintainer-overrides.csv",
              encoding="utf-8") as f:
        bad = [login for r in csv.DictReader(f)
               if (login := (r.get("login") or "").strip().lower())
               and login not in bf_logins]
    out.append(("eligibility/maintainer-overrides.csv", not bad,
                f"{len(bad)} login(s) not a bus-factor maintainer: " + ", ".join(bad[:3]) if bad
                else "every login carries an in-scope repo"))

    canon = canonical_repo_map()
    bad = []
    with open(ROOT / "data" / "risk" / "cve-package-overrides.csv",
              encoding="utf-8") as f:
        for r in csv.DictReader(f):
            slug = (r.get("github_repo") or "").strip().lower()
            if slug and canon.get(slug, slug) not in scope_slugs:
                bad.append(slug)
    bad = sorted(set(bad))
    out.append(("risk/cve-package-overrides.csv", not bad,
                f"{len(bad)} out-of-scope repo(s): " + ", ".join(bad[:3])
                + (" …" if len(bad) > 3 else "") if bad
                else "every override repo is in scope"))
    return out


def check_npm_downloads_status() -> list[Result]:
    """The npm downloads sidecar is well-formed and consistent with the data.

    Invariants the writer guarantees today (fetch_npm_data stamps a package
    only after its year rows are written): status ∈ {ok, not_found};
    checked_at parses; every stamped package has all model years in
    downloads.csv; a not_found package's years are all zero. Full coverage
    (every downloads.csv package stamped) is NOT asserted — legacy rows
    predate the sidecar and are only stamped as they are re-fetched.
    """
    from src.common.params import YEARS

    status_path = ROOT / "data" / "sources" / "npm" / "raw" / "downloads.status.csv"
    dl_path = ROOT / "data" / "sources" / "npm" / "raw" / "downloads.csv"
    if not status_path.exists():
        return [("npm downloads.status", False, "sidecar missing")]
    years = {str(y) for y in YEARS}
    dl: dict[str, dict[str, str]] = {}
    with open(dl_path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            dl.setdefault(r.get("package", ""), {})[r.get("year", "")] = \
                (r.get("downloads") or "").strip()
    problems: list[str] = []
    n = 0
    with open(status_path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            n += 1
            pkg = (r.get("package") or "").strip()
            status = (r.get("status") or "").strip()
            if status not in ("ok", "not_found"):
                problems.append(f"{pkg}: bad status {status!r}")
                continue
            try:
                datetime.strptime((r.get("checked_at") or "").strip(),
                                  "%Y-%m-%d %H:%M:%S.%f")
            except ValueError:
                problems.append(f"{pkg}: bad checked_at")
                continue
            pkg_years = dl.get(pkg, {})
            if not years <= set(pkg_years):
                problems.append(f"{pkg}: stamped but years missing from downloads.csv")
            elif status == "not_found" and any(pkg_years[y] not in ("", "0") for y in years):
                problems.append(f"{pkg}: not_found but has nonzero downloads")
    return [("npm downloads.status", not problems,
             f"{len(problems)} inconsistent row(s): " + "; ".join(problems[:3])
             + (" …" if len(problems) > 3 else "") if problems
             else f"{n} stamped packages consistent with downloads.csv")]


def check_value_domains() -> list[Result]:
    """value.csv enum columns stay inside their declared domains.

    platform ∈ {github, gitlab, bitbucket, sourcehut, codeberg, custom, ''}
    (docs/value.md); git_valid is strictly True/False (CLAUDE.md contract) —
    a stray value here means a writer regression, and every downstream
    filter compares against these exact strings.
    """
    platforms = {"github", "gitlab", "bitbucket", "sourcehut", "codeberg", "custom", ""}
    bad_platform: Counter = Counter()
    bad_valid: Counter = Counter()
    with open(ROOT / "data" / "value" / "value.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            p = (r.get("platform") or "").strip()
            if p not in platforms:
                bad_platform[p] += 1
            v = r.get("git_valid") or ""
            if v not in ("True", "False"):
                bad_valid[v] += 1
    out: list[Result] = []
    out.append(("value.csv:platform", not bad_platform,
                "all rows in the declared platform set" if not bad_platform
                else f"stray platform value(s): {dict(bad_platform)}"))
    out.append(("value.csv:git_valid", not bad_valid,
                "strictly True/False" if not bad_valid
                else f"non-boolean value(s): {dict(bad_valid)}"))
    return out


CHECKS = [check_dimension_csvs, check_risk_data, check_eligibility_data,
          check_value_data, check_value_criticality, check_validation_data,
          check_stats_data, check_preview_data,
          check_source_repo_id_integrity, check_source_value_id_agreement,
          check_source_schema_contract,
          check_score_component_coverage,
          check_score_input_completeness, check_long_format_keys,
          check_gitlab_host_nicknames, check_repo_id_shape,
          check_fetch_date_sanity, check_overrides_integrity,
          check_npm_downloads_status, check_value_domains]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--strict", action="store_true",
                        help="Exit non-zero if any check fails")
    args = parser.parse_args()

    console.print("\n[bold]Pipeline health check[/bold]\n")
    results: list[Result] = []
    for check in CHECKS:
        results.extend(check())

    table = Table(show_header=True, header_style="bold dim", padding=(0, 1))
    table.add_column("Check", style="bold")
    table.add_column("Status", justify="center")
    table.add_column("Detail", style="dim")
    failures = 0
    for label, ok, detail in results:
        if not ok:
            failures += 1
        table.add_row(label, "[green]OK[/green]" if ok else "[red]FAIL[/red]", detail)
    console.print(table)

    if failures:
        console.print(f"\n[red]{failures} check(s) failed[/red] of {len(results)}.\n")
    else:
        console.print(f"\n[green]All {len(results)} checks passed — pipeline consistent.[/green]\n")

    return 1 if (args.strict and failures) else 0


if __name__ == "__main__":
    sys.exit(main())
