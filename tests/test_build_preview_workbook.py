"""Tests for src/preview/build_preview_workbook.py — repos -> preview.xlsx."""
import csv

from openpyxl import load_workbook

from src.preview import build_preview_workbook as bpw

_STATS_MD = """# Pipeline Statistics

Some prose that must be skipped.

## Value

### Repo identity coverage

| Step | A | Total | Comment |
|---|--:|--:|---|
| Packages | 3,425 | 17,647 | package universe |
| **Valid repos** | **947** | **11,378** | upstream resolves |

more prose

### Coverage

| Signal | Filled | Pct |
|---|--:|--:|
| `openssf_crit` | 921 | 97.8% |
"""


def _write_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def _patch_stats_md(monkeypatch):
    """Feed the fixture markdown in place of the live generator."""
    monkeypatch.setattr(bpw, "_stats_markdown", lambda: _STATS_MD)


def test_cell_value_coerces_numbers_and_keeps_ids_as_text():
    assert bpw._cell_value("") is None
    assert bpw._cell_value("1") == 1
    assert bpw._cell_value("88.00") == 88.0
    assert bpw._cell_value("gh/61137153") == "gh/61137153"
    assert bpw._cell_value("True") == "True"


def test_build_writes_the_three_named_sheets_with_styled_filtered_headers(tmp_path, monkeypatch):
    repos_csv = tmp_path / "repos.csv"
    out = tmp_path / "preview.xlsx"
    _write_csv(repos_csv, ["repo", "risk_score", "repo_id"],
               [["a/keep", "88.00", "gh/1"], ["b/drop", "77.00", "gh/2"]])
    monkeypatch.setattr(bpw, "SHEETS", [("repos", repos_csv)])
    _patch_stats_md(monkeypatch)
    monkeypatch.setattr(bpw, "OUTPUT_FILE", out)

    bpw.build()

    wb = load_workbook(out)
    assert wb.sheetnames == bpw.SHEET_ORDER == ["repos", "components", "pipeline"]

    ws = wb["repos"]
    assert ws.max_row == 3          # header + 2 data rows
    assert ws.max_column == 2       # repo_id dropped from the repos sheet
    # header styling: bold white font on a filled (non-default) background.
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb == "001F3864"
    # data rows are NOT styled like the header.
    data_cell = ws.cell(row=2, column=1)
    assert data_cell.font.bold is not True
    # numeric column written as a real number, not text.
    assert ws.cell(row=2, column=2).value == 88.0
    # AutoFilter covers the used range after the drop; header row is frozen.
    assert ws.auto_filter.ref == "A1:B3"
    assert ws.freeze_panes == "A2"

    # a name in SHEETS that is not in SHEET_ORDER is never written
    monkeypatch.setattr(bpw, "SHEETS", [("repos", repos_csv), ("people", repos_csv)])
    bpw.build()
    assert load_workbook(out).sheetnames == ["repos", "components", "pipeline"]


def test_components_sheet_renders_methodology_tables(tmp_path, monkeypatch):
    """The components sheet: one colored banner per stage table, bold name
    cells, wrapped descriptions, and the weight tail formatted from params."""
    repos_csv = tmp_path / "repos.csv"
    out = tmp_path / "preview.xlsx"
    _write_csv(repos_csv, ["repo"], [["a/b"]])
    monkeypatch.setattr(bpw, "SHEETS", [("repos", repos_csv)])
    _patch_stats_md(monkeypatch)
    monkeypatch.setattr(bpw, "OUTPUT_FILE", out)

    bpw.build()

    # people.csv is a standalone CSV deliverable — not shipped in the workbook.
    assert [n for n, _ in bpw.SHEETS] == ["repos"]

    ws = load_workbook(out, rich_text=True)["components"]
    assert ws.sheet_view.showGridLines is False
    # value banner at B2: bold white on the stage green, boxed, merged B:C.
    banner = ws.cell(row=2, column=2)
    assert banner.value == "Value Components"
    assert banner.font.bold is True and banner.font.color.rgb == "00FFFFFF"
    assert banner.fill.start_color.rgb == "009BBB59"
    assert "B2:C2" in {str(r) for r in ws.merged_cells.ranges}
    # first row under it: bold name + wrapped description.
    name = ws.cell(row=3, column=2)
    assert name.value == "value_score" and name.font.bold is True
    desc = ws.cell(row=3, column=3)
    assert desc.alignment.wrap_text is True
    assert "pro-rata weighted blend" in str(desc.value)
    assert ws.row_dimensions[3].height > 15   # sized for the wrapped prose
    # weighted component carries the settings.json weight as a bold tail.
    openssf_desc = str(ws.cell(row=4, column=3).value)
    assert openssf_desc.endswith("Weight = 60%")
    # all four banners present, in stage order.
    banners = [c.value for c in ws["B"] if c.font and c.font.color
               and c.font.color.rgb == "00FFFFFF" and c.value]
    assert banners == ["Value Components", "Risk Components",
                       "Eligibility Components", "Preview Results"]


def test_build_skips_missing_csv_without_error(tmp_path, monkeypatch):
    repos_csv = tmp_path / "repos.csv"
    missing_csv = tmp_path / "does-not-exist.csv"
    out = tmp_path / "preview.xlsx"
    _write_csv(repos_csv, ["repo"], [["a/b"]])
    monkeypatch.setattr(bpw, "SHEETS", [("repos", missing_csv)])
    _patch_stats_md(monkeypatch)
    monkeypatch.setattr(bpw, "OUTPUT_FILE", out)

    bpw.build()

    wb = load_workbook(out)
    assert wb.sheetnames == ["repos", "components", "pipeline"]
    assert wb["repos"].max_row == 1    # empty sheet, no header written


def test_repo_url_per_platform():
    assert bpw._repo_url("a/b", "gh/123") == "https://github.com/a/b"
    assert bpw._repo_url("a/b", "gl/456") == "https://gitlab.com/a/b"
    assert (bpw._repo_url("mpc/mpc", "gl/inria-22470")
            == "https://gitlab.inria.fr/mpc/mpc")
    assert bpw._repo_url("a/b", "gl/unknown-1") is None
    assert bpw._repo_url("a/b", "") is None


def test_repos_sheet_hyperlinks_and_reviewed_decoration(tmp_path, monkeypatch):
    """Pins the reviewed repos-sheet design: hyperlinks, the single static
    fill (eligible purple), bold+centered score cells, the per-column-group
    conditional formats, and the fixed column widths."""
    repos_csv = tmp_path / "repos.csv"
    out = tmp_path / "preview.xlsx"
    header = ["repo", "value_score", "risk_score", "score",
              "oss", "eligible", "priority", "repo_id"]
    #          A       B              C             D        E      F           G           H
    _write_csv(repos_csv, header,
               [["a/keep", "70.25", "80.00", "90.00", "True", "True", "P1", "gh/1"],
                ["c/lab", "60.00", "70.00", "80.00", "False", "False", "", "gl/debian-9"]])
    monkeypatch.setattr(bpw, "SHEETS", [("repos", repos_csv)])
    _patch_stats_md(monkeypatch)
    monkeypatch.setattr(bpw, "OUTPUT_FILE", out)

    bpw.build()

    ws = load_workbook(out)["repos"]
    # repo hyperlinks still work, host derived from repo_id.
    assert ws.cell(row=2, column=1).hyperlink.target == "https://github.com/a/keep"
    assert ws.cell(row=3, column=1).hyperlink.target == "https://salsa.debian.org/c/lab"

    # eligible keeps its static purple fill; the score columns carry NO
    # static fill (their color comes from conditional scales).
    assert ws.cell(row=2, column=6).fill.start_color.rgb == "00E4DFEC"  # eligible
    assert ws.cell(row=2, column=2).fill.fill_type is None              # value_score
    assert ws.cell(row=2, column=3).fill.fill_type is None              # risk_score
    assert ws.cell(row=2, column=4).fill.fill_type is None              # score

    # value_score/risk_score data cells are bold and centered; repo is not.
    for col in (2, 3):
        assert ws.cell(row=2, column=col).font.bold is True
        assert ws.cell(row=2, column=col).alignment.horizontal == "center"
    assert ws.cell(row=2, column=1).alignment.horizontal != "center"    # repo

    # score cells display rounded ("0" format) but STORE the 2-decimal value.
    assert ws.cell(row=2, column=2).number_format == "0"
    assert ws.cell(row=2, column=2).value == 70.25
    assert ws.cell(row=2, column=1).number_format != "0"                # repo untouched

    # repo_id fuels the hyperlinks above, then leaves the sheet.
    final_headers = [c.value for c in ws[1]]
    assert "repo_id" not in final_headers
    assert ws.auto_filter.ref == ws.dimensions

    # conditional formats, one entry per column data range.
    cfs = {str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting}

    value_scale = cfs["B2:B3"][0].colorScale       # value_score: 3-color
    assert [c.rgb for c in value_scale.color] == \
        ["00F8696B", "00FFEB84", "0063BE7B"]       # red → yellow → GREEN max
    assert [(v.type, v.val) for v in value_scale.cfvo] == \
        [("num", 0.0), ("num", 50.0), ("num", 100.0)]

    risk_scale = cfs["C2:C3"][0].colorScale        # risk_score: reversed
    assert [c.rgb for c in risk_scale.color] == \
        ["0063BE7B", "00FFEB84", "00F8696B"]       # green → yellow → RED max

    score_scale = cfs["D2:D3"][0].colorScale       # score: 2-color to blue
    assert [c.rgb for c in score_scale.color] == ["00FFFFFF", "002C96DE"]

    for rng in ("E2:E3", "F2:F3"):                 # oss + eligible booleans
        rules = {r.formula[0]: r for r in cfs[rng]}
        assert all(r.type == "cellIs" and r.operator == "equal"
                   for r in rules.values())
        assert rules['"True"'].dxf.font.color.rgb == "00006100"   # dark green
        assert rules['"False"'].dxf.font.color.rgb == "009C0006"  # dark red
        assert rules['"True"'].dxf.fill is None                   # font-only

    priority_rule = cfs["G2:G3"][0]                # priority: non-blank fill
    assert priority_rule.type == "expression"
    assert priority_rule.formula == ["NOT(ISBLANK(G2))"]
    assert priority_rule.dxf.fill.start_color.rgb == "00DDEBF7"

    # fixed column widths by header name.
    assert ws.column_dimensions["A"].width == 24   # repo
    assert ws.column_dimensions["G"].width == 8    # priority


def test_stats_sheet_renders_markdown_tables(tmp_path, monkeypatch):
    """Every preview pipeline-sheet table lands as a stacked block: bold section heading,
    styled header row, markdown dressing stripped, counts as real numbers."""
    repos_csv = tmp_path / "repos.csv"
    out = tmp_path / "preview.xlsx"
    _write_csv(repos_csv, ["repo_id"], [["gh/1"]])
    monkeypatch.setattr(bpw, "SHEETS", [("repos", repos_csv)])
    _patch_stats_md(monkeypatch)
    monkeypatch.setattr(bpw, "OUTPUT_FILE", out)

    bpw.build()

    ws = load_workbook(out)["pipeline"]
    assert ws.sheet_view.showGridLines is False
    got = [[c.value for c in row] for row in ws.iter_rows()]
    # column A is an empty gutter — every block starts at column B; row 1 blank
    assert all(row[0] is None for row in got)
    assert got[0] == [None] * len(got[0])
    # '## Value' renders as a stage BANNER: bold on the header fill, row 2
    assert got[1][1] == "Value"
    banner = ws.cell(row=2, column=2)
    assert banner.font.bold is True and banner.fill.start_color.rgb == "001F3864"
    # two blank rows, then block 1: bold sub-title directly above the header
    assert got[4][1] == "Repo identity coverage"
    assert ws.cell(row=5, column=2).font.bold is True
    assert got[5][1:5] == ["Step", "A", "Total", "Comment"]
    assert ws.cell(row=6, column=2).fill.start_color.rgb == "001F3864"
    assert got[6][1:4] == ["Packages", 3425, 17647]
    assert got[7][1:4] == ["Valid repos", 947, 11378]     # **bold** stripped
    # bold row + typed numbers + right-aligned numeric headers + box border
    assert ws.cell(row=8, column=2).font.bold is True
    assert ws.cell(row=7, column=3).number_format == "#,##0"
    assert ws.cell(row=6, column=3).alignment.horizontal == "right"   # "A"
    assert ws.cell(row=6, column=2).alignment.horizontal != "right"   # "Step"
    assert ws.cell(row=6, column=2).border.left.style == "thin"
    assert ws.cell(row=8, column=2).border.bottom.style == "thin"
    # two blank rows, then block 2 under its own sub-title
    assert got[8] == [None] * len(got[8])
    assert got[9] == [None] * len(got[9])
    assert got[10][1] == "Coverage"
    assert got[12][1:3] == ["openssf_crit", 921]
    # percent cells become real fractions with a percent format
    pct = ws.cell(row=13, column=4)
    assert pct.value == 0.978 and pct.number_format == "0.0%"
