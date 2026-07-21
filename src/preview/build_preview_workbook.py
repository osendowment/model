#!/usr/bin/env python3
"""Build data/preview/preview.xlsx — repos.csv + methodology + stats as one workbook.

The terminal step of the pipeline, a spreadsheet for non-technical review.
Sheet 1, 'repos', is src.preview.build_results -> repos.csv; numeric-looking cells
(scores, counts) are written as real Excel numbers, not text, so
sorting/filtering behaves numerically rather than alphabetically.
(data.csv stays a standalone CSV deliverable — it is not shipped in the
workbook.)

Sheet 2, 'components', documents the methodology: one banner-headed
table per stage (Value / Risk / Eligibility / Preview Results), one row
per score or component — name + a prose description of its data sources
and formula. The value-component weights are formatted at build time from
`settings.json` (via src.common.params) so the text cannot drift from the
config.

Sheet 3, 'pipeline', renders every pipeline count/funnel/coverage table
as stacked blocks — each under its section heading, with the same styled
header row. The tables come straight from the generator
(`scripts/stats.py`, its markdown renderer), computed from the live CSVs at
build time: this sheet IS the single home of the pipeline's numbers.

The repos sheet gets:
    - a styled header row (bold white text on a dark-blue fill) so it reads
      as a header at a glance, distinct from the data rows
    - an AutoFilter over the full used range, so every column has a
      dropdown filter/sort control
    - a frozen header row (`freeze_panes`), so the header (and its filter
      dropdowns) stays visible while scrolling a long sheet

The repos sheet is additionally decorated (matching the reviewed design):
each `repo` cell hyperlinks to the repo's home page (host derived from
`repo_id`); the value columns carry a red→yellow→green 3-color scale
anchored at 0/50/100 and the risk columns the reversed scale (higher risk
= red); `score` fades white→blue; the boolean columns render True/False as
dark-green/dark-red text, with `eligible` also on a static light-purple
column fill; non-empty `priority` / `eco_priority` cells fill light blue. Numeric, boolean
and priority data cells are centered, `value_score`/`risk_score` are bold,
and every known column gets a fixed width.

Usage:
    uv run python -m src.preview.build_preview_workbook
"""

import csv
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

from src.common.params import (
    VALUE_SCORE_CENTRALITY_WEIGHT,
    VALUE_SCORE_CRIT_WEIGHT,
    VALUE_SCORE_ECO_CRIT_WEIGHT,
    VALUE_SCORE_PR_WEIGHT,
)

console = Console()

ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = ROOT / "data"
REPOS_CSV = DATA_DIR / "preview" / "repos.csv"
STATS_SCRIPT = ROOT / "scripts" / "stats.py"
OUTPUT_FILE = DATA_DIR / "preview" / "preview.xlsx"

SHEETS = [("repos", REPOS_CSV)]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=13)

def _fill(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

# repos sheet decoration (the reviewed design). Column groups are keyed by
# header NAME so the layout survives column reorders; unknown headers are
# simply skipped.
REPOS_STATIC_FILLS = {"eligible": _fill("E4DFEC")}  # light purple

# All four value_score components + the blend itself share the value scale.
REPOS_VALUE_SCALE_COLS = ["top_eco_pct", "pr_score", "openssf_crit", "eco_crit",
                          "value_score"]
REPOS_RISK_SCALE_COLS = ["concentration", "complexity", "security", "workload",
                         "risk_score"]
REPOS_SCORE_COL = "score"
REPOS_BOOL_COLS = ["oss", "intent", "nonprofit", "active", "eligible"]
REPOS_PRIORITY_COLS = ["priority", "eco_priority"]
REPOS_BOLD_COLS = ["value_score", "risk_score"]
# Score cells DISPLAY as whole numbers but keep their full 2-decimal value
# (visible in the formula bar / on export) — pure number formatting.
REPOS_ROUNDED_COLS = (REPOS_VALUE_SCALE_COLS + REPOS_RISK_SCALE_COLS
                      + [REPOS_SCORE_COL])
ROUNDED_FORMAT = "0"
REPOS_CENTERED_COLS = (REPOS_VALUE_SCALE_COLS + REPOS_RISK_SCALE_COLS
                       + [REPOS_SCORE_COL] + REPOS_BOOL_COLS
                       + REPOS_PRIORITY_COLS)

# Reviewed widths (from the hand-tuned workbook). Columns not listed —
# pr_score, openssf_crit, the four risk dimensions, risk_score — stay at
# Excel's default width.
REPOS_COLUMN_WIDTHS = {
    "repo": 24, "platform": 9, "ecosystem": 10, "top_eco_pkg": 16,
    "top_eco_pct": 10, "eco_crit": 8, "value_score": 10,
    "score": 9, "oss": 7, "intent": 8, "nonprofit": 9,
    "active": 8, "eligible": 9, "priority": 8, "eco_priority": 12,
}

# repos.csv keeps repo_id as its join/trace key, but the sheet is for human
# readers: the column fuels the repo hyperlinks, then leaves the sheet.
REPOS_DROP_COLS = ["repo_id"]

# 3-color scale anchors (Excel's classic red/yellow/green) + the score blue.
SCALE_RED, SCALE_YELLOW, SCALE_GREEN = "F8696B", "FFEB84", "63BE7B"
SCORE_BLUE = "2C96DE"
BOOL_TRUE_FONT = Font(color="006100")    # dark green
BOOL_FALSE_FONT = Font(color="9C0006")   # dark red
PRIORITY_FILL = _fill("DDEBF7")          # light blue
CENTER = Alignment(horizontal="center")
BOLD = Font(bold=True)


def _cell_value(raw: str) -> int | float | str | None:
    """Coerce a CSV string to a real Excel number where possible.

    Blank -> None (an empty cell, not the literal string ""). Otherwise try
    int, then float, else keep the original string (ids like `gh/12345`,
    booleans like `True`, free text).
    """
    if raw == "":
        return None
    try:
        return int(raw)
    except ValueError:
        pass
    try:
        return float(raw)
    except ValueError:
        return raw


def _write_sheet(ws: Worksheet, csv_path: Path) -> int:
    """Write one CSV's contents into `ws`, styling row 1 as the header.
    Returns the number of data rows written (0 if the CSV is missing/empty).
    """
    if not csv_path.exists():
        return 0
    with open(csv_path, encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        return 0

    header, *data_rows = rows
    ws.append(header)
    for row in data_rows:
        ws.append([_cell_value(v) for v in row])

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return len(data_rows)


def _repo_url(repo: str, repo_id: str) -> str | None:
    """The repo's home page, derived from its platform-qualified id:
    `gh/<n>` → github.com, bare `gl/<n>` → gitlab.com, `gl/<nickname>-<n>` →
    that self-hosted GitLab instance (nickname resolved via
    gitlab_client.HOST_NICKNAMES). Unknown/blank id → no link."""
    if repo_id.startswith("gh/"):
        return f"https://github.com/{repo}"
    if repo_id.startswith("gl/"):
        from src.sources.gitlab.gitlab_client import HOST_NICKNAMES
        rest = repo_id[3:]
        if rest.isdigit():
            return f"https://gitlab.com/{repo}"
        nickname = rest.rsplit("-", 1)[0]
        host = {n: h for h, n in HOST_NICKNAMES.items() if n}.get(nickname)
        return f"https://{host}/{repo}" if host else None
    return None


def _decorate_repos_sheet(ws: Worksheet) -> None:
    """repos-only dressing (the reviewed design): hyperlink each `repo` cell
    to the repo's home page (host from `repo_id`), apply the static fills,
    center/bold the metric cells and display scores rounded (full 2-decimal
    values stay stored), drop the REPOS_DROP_COLS join keys from the sheet,
    add the conditional formats per column group, and set the fixed column
    widths."""
    headers = {c.value: c.column for c in ws[1]}  # name -> 1-based col index
    repo_col, id_col = headers.get("repo"), headers.get("repo_id")
    fill_cols = {headers[n]: f for n, f in REPOS_STATIC_FILLS.items() if n in headers}
    center_cols = {headers[n] for n in REPOS_CENTERED_COLS if n in headers}
    bold_cols = {headers[n] for n in REPOS_BOLD_COLS if n in headers}
    rounded_cols = {headers[n] for n in REPOS_ROUNDED_COLS if n in headers}
    for row in ws.iter_rows(min_row=2):
        if repo_col and id_col:
            cell = row[repo_col - 1]
            url = _repo_url(str(cell.value or ""), str(row[id_col - 1].value or ""))
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"
        for col, fill in fill_cols.items():
            row[col - 1].fill = fill
        for col in center_cols:
            row[col - 1].alignment = CENTER
        for col in bold_cols:
            row[col - 1].font = BOLD
        for col in rounded_cols:
            row[col - 1].number_format = ROUNDED_FORMAT
    # Highest index first so earlier deletions don't shift later ones.
    for col in sorted((headers[n] for n in REPOS_DROP_COLS if n in headers),
                      reverse=True):
        ws.delete_cols(col)
    headers = {c.value: c.column for c in ws[1]}
    ws.auto_filter.ref = ws.dimensions
    _add_repos_conditional_formats(ws, headers)
    for name, width in REPOS_COLUMN_WIDTHS.items():
        if name in headers:
            ws.column_dimensions[get_column_letter(headers[name])].width = width


def _add_repos_conditional_formats(ws: Worksheet, headers: dict) -> None:
    """Conditional formats on the repos sheet, one per column group, each
    over the column's data range (row 2..max_row):

      - value columns: 3-color scale, 0=red → 50=yellow → 100=green;
      - risk columns: the same scale reversed — higher risk reads red;
      - `score`: 2-color scale, 0=white → 100=blue;
      - boolean columns: "True" in dark-green text, "False" in dark-red
        (font only, so `eligible`'s static purple fill shows through);
      - `priority`: light-blue fill on non-empty cells.
    """
    def data_range(name: str) -> tuple[str, str] | tuple[None, None]:
        col = headers.get(name)
        if not col or ws.max_row < 2:
            return None, None
        letter = get_column_letter(col)
        return letter, f"{letter}2:{letter}{ws.max_row}"

    for name in REPOS_VALUE_SCALE_COLS:
        _, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color=SCALE_RED,
                mid_type="num", mid_value=50, mid_color=SCALE_YELLOW,
                end_type="num", end_value=100, end_color=SCALE_GREEN))
    for name in REPOS_RISK_SCALE_COLS:
        _, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color=SCALE_GREEN,
                mid_type="num", mid_value=50, mid_color=SCALE_YELLOW,
                end_type="num", end_value=100, end_color=SCALE_RED))
    _, rng = data_range(REPOS_SCORE_COL)
    if rng:
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            end_type="num", end_value=100, end_color=SCORE_BLUE))
    for name in REPOS_BOOL_COLS:
        _, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"True"'], font=BOOL_TRUE_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"False"'], font=BOOL_FALSE_FONT))
    for name in REPOS_PRIORITY_COLS:
        letter, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f"NOT(ISBLANK({letter}2))"], fill=PRIORITY_FILL))


def _md_cell(raw: str) -> tuple[int | float | str | None, str | None]:
    """One markdown table cell → (Excel value, number format).

    Strips the markdown dressing (`**bold**`, `` `code` `` — including code
    spans inside prose), then coerces to REAL typed cells so Excel never
    flags "number stored as text": thousands-separated counts become ints
    (format `#,##0`), decimals become floats (`0.00`), and percent strings
    ("97.8%") become fractions with a percent format (`0.0%`, or `0%` when
    the source had no decimals). Prose stays text (no format); a
    blank/placeholder cell becomes an empty cell."""
    s = raw.strip().strip("*").replace("`", "").strip()
    if s in ("", "—", "·"):
        return None, None
    if s.endswith("%"):
        num = s[:-1].replace(",", "").strip()
        try:
            return float(num) / 100.0, ("0.0%" if "." in num else "0%")
        except ValueError:
            return s, None
    plain = s.replace(",", "")
    try:
        return int(plain), "#,##0"
    except ValueError:
        pass
    try:
        return float(plain), "0.00"
    except ValueError:
        return s, None


def _is_md_separator(line: str) -> bool:
    """A markdown table's |---|--:|---| alignment row (no letters/digits)."""
    body = line.strip().strip("|")
    return bool(body) and set(body) <= set("-:| ")


def _stats_markdown() -> str:
    """The stats tables as markdown, straight from the generator.

    `scripts/stats.py` computes every count from the live CSVs; its
    `markdown()` renderer is the one source of truth for the numbers on
    this sheet (the preview pipeline sheet no longer exists)."""
    import importlib.util
    spec = importlib.util.spec_from_file_location("_stats_gen", STATS_SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.markdown(mod.value_stats(), mod.risk_stats(), mod.eligibility_stats())


THIN = Side(style="thin", color="000000")


def _with_sides(cell, **sides) -> None:
    """Merge border sides onto a cell without clobbering the ones it has."""
    b = cell.border
    cell.border = Border(left=sides.get("left", b.left),
                         right=sides.get("right", b.right),
                         top=sides.get("top", b.top),
                         bottom=sides.get("bottom", b.bottom))


def _render_table(ws: Worksheet, block: list[list[str]],
                  top_row: int, start_col: int) -> int:
    """Render one parsed markdown table at (top_row, start_col); returns its
    last row. Handles typed cells + number formats, whole-row bold from
    `**` markers, `^` separator lines, numeric-header right-alignment and
    the thin box outline."""
    n_cols = max(len(cells) for cells in block)
    numeric_col = [True] * n_cols
    has_data_col = [False] * n_cols
    separator_rows: list[int] = []
    for j, cells in enumerate(block):
        row_idx = top_row + j
        first = cells[0].strip()
        if first.startswith("^"):
            separator_rows.append(row_idx)
            cells = [first.lstrip("^")] + cells[1:]
        bold = cells[0].strip().startswith("**")
        for k, (v, fmt) in enumerate(_md_cell(x) for x in cells):
            cell = ws.cell(row=row_idx, column=start_col + k, value=v)
            if fmt:
                cell.number_format = fmt
            if j == 0:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                if bold:
                    cell.font = Font(bold=True)
                if v is not None:
                    has_data_col[k] = True
                    if not isinstance(v, (int, float)):
                        numeric_col[k] = False
    for k in range(n_cols):
        if has_data_col[k] and numeric_col[k]:
            ws.cell(row=top_row, column=start_col + k).alignment = \
                Alignment(horizontal="right")
    # thin box outline + separator lines above marked rows
    last_row = top_row + len(block) - 1
    last_col = start_col + n_cols - 1
    for row in range(top_row, last_row + 1):
        _with_sides(ws.cell(row=row, column=start_col), left=THIN)
        _with_sides(ws.cell(row=row, column=last_col), right=THIN)
    for col in range(start_col, last_col + 1):
        _with_sides(ws.cell(row=top_row, column=col), top=THIN)
        _with_sides(ws.cell(row=last_row, column=col), bottom=THIN)
        for row in separator_rows:
            _with_sides(ws.cell(row=row, column=col), top=THIN)
    return last_row


BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)

# Stage banner fills (matched to the reviewed sheet): value blue, risk
# orange, eligibility purple; anything else falls back to the header navy.
BANNER_FILLS = {
    "Stage 1": _fill("5B9BD5"),
    "Stage 2": _fill("ED7D31"),
    "Stage 3": _fill("8064A2"),
}
BANNER_SPAN_COLS = 13  # fill A..M so the banner reads as a full-width bar


def _write_pipeline_sheet(ws: Worksheet, md_text: str) -> int:
    """Render the generator's markdown as the pipeline sheet.

    Layout rules (mirroring the reviewed design):
      - gridlines off; column A is an empty gutter, content starts at B;
      - `## …` headings are STAGE BANNERS — bold white on the header fill;
      - `### …` headings are bold sub-titles, printed once above their table;
      - a `<!-- beside -->` directive renders the NEXT table to the right of
        the previous one (same rows, one gap column) — used for the CPP
        ecosystem breakdown next to the value funnel;
      - `^` first-cell markers draw separator lines, `**` rows are bold.
    """
    ws.sheet_view.showGridLines = False
    lines = md_text.splitlines()
    heading = ""
    heading_level = 3
    heading_emitted = True
    beside = False
    prev = None  # (top_row, start_col, n_cols, last_row) of the previous table
    cursor = 2   # next free row (row 1 stays blank, like the column-A gutter)
    tables = 0
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("#"):
            heading_level = len(line) - len(line.lstrip("#"))
            heading = line.lstrip("#").strip()
            if heading_level == 2:
                # Stage banner — standalone, emitted immediately, full-width.
                fill = next((f for k, f in BANNER_FILLS.items()
                             if heading.startswith(k)), HEADER_FILL)
                for col in range(1, BANNER_SPAN_COLS + 1):
                    ws.cell(row=cursor, column=col).fill = fill
                cell = ws.cell(row=cursor, column=2, value=heading)
                cell.font = BANNER_FONT
                cursor += 3  # banner + two blank rows
                heading_emitted = True
            else:
                heading_emitted = False
            i += 1
            continue
        if line == "<!-- beside -->":
            beside = True
            i += 1
            continue
        if not line.startswith("|"):
            i += 1
            continue
        block = []
        while i < len(lines) and lines[i].strip().startswith("|"):
            if not _is_md_separator(lines[i]):
                block.append([x for x in lines[i].strip().strip("|").split("|")])
            i += 1
        if not block:
            continue

        if beside and prev is not None:
            top_row, start_col = prev[0], prev[1] + prev[2] + 1
            last = _render_table(ws, block, top_row, start_col)
            cursor = max(cursor, last + 2)
            prev = (top_row, start_col, max(len(r) for r in block), last)
            beside = False
            tables += 1
            continue

        if not heading_emitted:
            cell = ws.cell(row=cursor, column=2, value=heading)
            cell.font = SECTION_FONT
            cursor += 1  # sub-title sits directly above its table
            heading_emitted = True
        top_row = cursor
        last = _render_table(ws, block, top_row, 2)
        prev = (top_row, 2, max(len(r) for r in block), last)
        cursor = last + 3  # two blank rows between blocks
        tables += 1

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 60  # metric labels carry the folded descriptions
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 26  # beside-table label column
    for col in "JKLM":
        ws.column_dimensions[col].width = 12
    return tables


# ---------------------------------------------------------------------------
# components sheet — methodology tables, one row per score/component.
# Format: a colored full-width banner per table (matching the stage hues used
# across the workbook), then `name | description` rows. Weights are formatted
# from settings.json so this text cannot drift from the config.

def _w(desc: str, weight: float) -> CellRichText:
    """Description + a bold 'Weight = N%' tail (rich text, single cell)."""
    return CellRichText(desc + " ", TextBlock(InlineFont(b=True),
                                              f"Weight = {weight:.0%}"))


COMPONENT_TABLES: list[tuple[str, str, list[tuple[str, object]]]] = [
    ("Value Components", "9BBB59", [
        ("value_score",
         "A 0-100 pro-rata weighted blend of the four components below "
         "(weights from settings.json). Only the components present for a "
         "repo are summed and renormalized by their weight total, so a repo "
         "missing one still lands on the same scale; at least 2 present "
         "components are required, else blank. Criticality-dominant by "
         "design so a foundational-but-quiet micro-dep can't outrank a "
         "genuinely critical project."),
        ("openssf_crit",
         _w("The OpenSSF Criticality Score tool run per repo (GitHub-only): "
            "blends activity signals — commit cadence, contributor count, "
            "org diversity, dependents, issue activity — into a 0-1 score, "
            "stored ×100. Source: data/sources/openssf/criticality.csv; "
            "non-empty for every valid class-A GitHub repo.",
            VALUE_SCORE_CRIT_WEIGHT)),
        ("eco_crit",
         _w("Whether ecosyste.ms lists the repo's package on its curated "
            "critical-infrastructure list: 100 = on the list, 0 = explicitly "
            "not, blank = unknown (never a fake 0). Covers GitHub AND "
            "GitLab, so it is the importance signal GitLab repos still get. "
            "Source: data/sources/ecosystems/criticality.csv.",
            VALUE_SCORE_ECO_CRIT_WEIGHT)),
        ("top_eco_pct",
         _w("The repo's PageRank position inside its strongest ecosystem "
            "(top_eco). Per registry (npm / PyPI / crates / cpp = Debian + "
            "Homebrew): download stats pick the top packages (95% of "
            "cumulative downloads), their dependency tree is fetched, and a "
            "download-personalized PageRank (α = 0.85) ranks every node; "
            "top_eco_pct = 100 − cumulative-PR percentile, higher = better.",
            VALUE_SCORE_CENTRALITY_WEIGHT)),
        ("pr_score",
         _w("Cross-ecosystem dependency MASS, complementing top_eco_pct's "
            "position. The repo's summed package PageRank per ecosystem is "
            "ln-scaled and min-max normalized; the per-eco values combine "
            "as a p = 2 norm (a real second-ecosystem footprint adds ~41%, "
            "a token listing ~nothing), rescaled so the top repo = 100.",
            VALUE_SCORE_PR_WEIGHT)),
    ]),
    ("Risk Components", "C0504D", [
        ("risk_score",
         "Geometric mean of the four dimension scores below, each 0-100 "
         "with higher = riskier. Percentile-based dimensions use "
         "worst-pinned CDF ranks over the risk scope (valid class-A repos, "
         "archived included), direction-oriented so 'worse' always ranks "
         "high."),
        ("concentration",
         "Contributor concentration from a treeless git clone's commit log "
         "(git log --no-merges, mailmap applied, identities merged, bots "
         "dropped), windowed to the last 5 complete years. Bus factor "
         "(contributors covering 50% of commits) and HHI (sum of squared "
         "commit shares) combine as sqrt((100/bf) × (hhi/100)) — absolute "
         "scales, not percentiles, so a one-person repo pins 100."),
        ("complexity",
         "Code size and complexity at the end-of-year-pinned SHA from one "
         "sparse checkout: scc measures lines of code, lizard the "
         "per-function McCabe cyclomatic maximum. Both are "
         "percentile-ranked and the score is their geometric mean — big "
         "AND gnarly ranks worse than either alone."),
        ("security",
         "Worst-of two axes: the OpenSSF Scorecard score (API, deps.dev "
         "fallback) percentile-ranked INVERTED, and a neutral-anchored CVE "
         "score from OSV — 0 known CVEs = 50 ('none known' is not proof of "
         "safety), ≥1 CVE ranked among non-zero repos into (50, 100]. "
         "Taking the max means real CVEs are never masked by a good "
         "Scorecard, and vice versa."),
        ("workload",
         "Maintenance burden per maintainer: LOC per active contributor, "
         "CVEs per active contributor, and net-new issues per active "
         "contributor (GitHub Issues Search + GitLab issues API, merged by "
         "repo_id, per-year opened − closed) — "
         "each percentile-ranked, combined by geometric mean. The divisor "
         "is the same 5-year active-contributor count concentration uses "
         "(AC = 0 scored as AC = 1)."),
    ]),
    ("Eligibility Components", "8064A2", [
        ("eligible",
         "Plain AND of the four boolean checks below — no weights. "
         "Ineligible repos stay in the table with the failing flag "
         "visible."),
        ("oss",
         "One SPDX license resolved per repo — manual override → registry "
         "license of the repo's packages → GitHub Licensee → GitLab API — "
         "then checked against the approved set: SPDX isOsiApproved ∪ "
         "(isFsfLibre − content licenses) ∪ curated extras, so an "
         "FSF-libre software licence the OSI never formally reviewed still "
         "counts, while font/doc licences do not. Expression-aware: 'mit OR "
         "apache-2.0' counts if any component is approved. Ternary: True / "
         "False / blank (unknown ≠ known-non-OSS)."),
        ("intent",
         "True when the repo shows ANY funding signal: GitHub Sponsors "
         "enabled, FUNDING.yml, funding.json, npm/PyPI funding field, a "
         "real Open Collective, a bus-factor maintainer with personal "
         "Sponsors, an institutional host/owner, or a curated PayPal. "
         "Propagates at the owner level: one repo declaring a channel "
         "flips intent = True for all that owner's repos."),
        ("nonprofit",
         "Defaults True; flips False only when a curated/scraped COMPANY "
         "host or owner backs the repo (foundation rosters + "
         "overrides.csv). Company-backed projects are already resourced — "
         "ineligible but kept visible."),
        ("active",
         "active = NOT eol AND NOT archived. eol is a manual per-repo "
         "verdict in overrides.csv informed by registry "
         "deprecation/yank/endoflife.date advisories; archived is the "
         "GitHub flag, with no exemption — a project whose canonical "
         "upstream is off GitHub is repointed there in value/overrides.csv "
         "(e.g. glibc → sourceware, pixman → gitlab.freedesktop.org)."),
    ]),
    ("Preview Results", "4F81BD", [
        ("score",
         "sqrt(value_score × risk_score) — unnormalized geometric mean on "
         "the same 0-100 scale as its inputs (top row ≈ 76). Both "
         "dimensions must be high; computed for every row with both scores "
         "present, regardless of eligibility."),
        ("priority",
         "Dense rank (1, 2, 3, …) by score descending over eligible rows "
         "only — blank for ineligible or unscored rows."),
        ("eco_priority",
         "The same rank restarted within each ecosystem, so npm rank 1 and "
         "PyPI rank 1 are both present. Same rows and same order as "
         "priority — only the grouping differs."),
    ]),
]

COMPONENTS_DESC_WIDTH = 116          # col C width ≈ chars per line (reviewed)
COMPONENTS_BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_components_sheet(ws: Worksheet) -> int:
    """Render COMPONENT_TABLES as stacked `name | description` tables, each
    under its colored banner. Returns the number of tables written."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = COMPONENTS_DESC_WIDTH
    row = 2
    for title, rgb, entries in COMPONENT_TABLES:
        banner = ws.cell(row=row, column=2, value=title)
        banner.font = BANNER_FONT
        for col in (2, 3):
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(rgb)
            cell.border = COMPONENTS_BOX
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
        for name, desc in entries:
            ncell = ws.cell(row=row, column=2, value=name)
            ncell.font = Font(bold=True)
            ncell.alignment = Alignment(vertical="center")
            ncell.border = COMPONENTS_BOX
            dcell = ws.cell(row=row, column=3, value=desc)
            dcell.alignment = Alignment(vertical="center", wrap_text=True)
            dcell.border = COMPONENTS_BOX
            # openpyxl can't auto-fit wrapped rows — estimate from length.
            lines = math.ceil(len(str(desc)) / (COMPONENTS_DESC_WIDTH - 10))
            ws.row_dimensions[row].height = max(15, lines * 14 + 6)
            row += 1
        row += 2  # two blank rows between tables
    return len(COMPONENT_TABLES)


# Tab order, left to right: the answer, then how it was reached, then how much
# of the world it covers.
SHEET_ORDER = ["repos", "components", "pipeline"]


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet; we name our own
    csv_sheets = dict(SHEETS)
    decorate = {"repos": _decorate_repos_sheet}

    for name in SHEET_ORDER:
        if name not in csv_sheets:
            continue
        ws = wb.create_sheet(name)
        n = _write_sheet(ws, csv_sheets[name])
        if n:
            decorate[name](ws)
        console.print(f"  [cyan]{name}[/cyan]: {n:,} rows ← {csv_sheets[name]}")

    ws = wb.create_sheet("components", SHEET_ORDER.index("components"))
    n = _write_components_sheet(ws)
    console.print(f"  [cyan]components[/cyan]: {n} methodology tables (static)")

    ws = wb.create_sheet("pipeline", SHEET_ORDER.index("pipeline"))
    n = _write_pipeline_sheet(ws, _stats_markdown())
    console.print(f"  [cyan]pipeline[/cyan]: {n} tables ← scripts/stats.py (live CSVs)")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)


def main() -> None:
    console.print("[bold]Building preview.xlsx...[/bold]\n")
    build()
    console.print(f"\n[dim]Wrote {OUTPUT_FILE}[/dim]")


if __name__ == "__main__":
    main()
