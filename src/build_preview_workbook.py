#!/usr/bin/env python3
"""Build data/preview/preview.xlsx — repos.csv + people.csv + stats as one workbook.

The terminal step of the pipeline: combines the two preview CSVs
(src.build_results -> repos.csv, src.build_people -> people.csv) into a
single spreadsheet for non-technical review, one sheet per CSV ('repos',
'people'). Numeric-looking cells (scores, counts, ids) are written as real
Excel numbers, not text, so sorting/filtering behaves numerically rather
than alphabetically.

A third sheet, 'stats', renders every pipeline count/funnel/coverage table
as stacked blocks — each under its section heading, with the same styled
header row. The tables come straight from the generator
(`scripts/stats.py`, its markdown renderer), computed from the live CSVs at
build time: this sheet IS the single home of the pipeline's numbers.

The CSV sheets get:
    - a styled header row (bold white text on a dark-blue fill) so it reads
      as a header at a glance, distinct from the data rows
    - an AutoFilter over the full used range, so every column has a
      dropdown filter/sort control
    - a frozen header row (`freeze_panes`), so the header (and its filter
      dropdowns) stays visible while scrolling a long sheet

The repos sheet is additionally decorated (matching the reviewed design):
each `repo` cell hyperlinks to the repo's home page (host derived from
`repo_id`); the value columns carry a red→yellow→green 3-color scale
anchored at 0/50/100 and the risk columns the reversed scale (higher risk
= red); `score` fades white→blue; the boolean columns render True/False as
dark-green/dark-red text, with `eligible` also on a static light-purple
column fill; non-empty `priority` cells fill light blue. Numeric, boolean
and priority data cells are centered, `value_score`/`risk_score` are bold,
and every known column gets a fixed width.

Usage:
    uv run python -m src.build_preview_workbook
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

console = Console()

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
REPOS_CSV = DATA_DIR / "preview" / "repos.csv"
PEOPLE_CSV = DATA_DIR / "preview" / "people.csv"
STATS_SCRIPT = ROOT / "scripts" / "stats.py"
OUTPUT_FILE = DATA_DIR / "preview" / "preview.xlsx"

SHEETS = [("repos", REPOS_CSV), ("people", PEOPLE_CSV)]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=13)

def _fill(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

# repos sheet decoration (the reviewed design). Column groups are keyed by
# header NAME so the layout survives column reorders; unknown headers are
# simply skipped.
REPOS_STATIC_FILLS = {"eligible": _fill("E4DFEC")}  # light purple

REPOS_VALUE_SCALE_COLS = ["openssf_crit", "eco_crit", "top_eco_pct", "value_score"]
REPOS_RISK_SCALE_COLS = ["concentration", "complexity", "security", "workload",
                         "risk_score"]
REPOS_SCORE_COL = "score"
REPOS_BOOL_COLS = ["oss", "intent", "nonprofit", "active", "eligible"]
REPOS_PRIORITY_COL = "priority"
REPOS_BOLD_COLS = ["value_score", "risk_score"]
REPOS_CENTERED_COLS = (REPOS_VALUE_SCALE_COLS + REPOS_RISK_SCALE_COLS
                       + [REPOS_SCORE_COL] + REPOS_BOOL_COLS
                       + [REPOS_PRIORITY_COL])

REPOS_COLUMN_WIDTHS = {
    "repo": 24, "language": 11, "ecosystem": 10, "top_eco_pkg": 16,
    "openssf_crit": 10, "eco_crit": 8, "top_eco_pct": 10, "value_score": 10,
    "concentration": 10, "complexity": 10, "security": 10, "workload": 10,
    "risk_score": 10, "score": 9, "oss": 7, "intent": 8, "nonprofit": 9,
    "active": 8, "eligible": 9, "priority": 8, "repo_id": 13,
}

# 3-color scale anchors (Excel's classic red/yellow/green) + the score blue.
SCALE_RED, SCALE_YELLOW, SCALE_GREEN = "F8696B", "FFEB84", "63BE7B"
SCORE_BLUE = "2C96DE"
BOOL_TRUE_FONT = Font(color="006100")    # dark green
BOOL_FALSE_FONT = Font(color="9C0006")   # dark red
PRIORITY_FILL = _fill("DDEBF7")          # light blue
CENTER = Alignment(horizontal="center")
BOLD = Font(bold=True)


def _cell_value(raw: str) -> int | float | str | None:
    """Coerce a CSV string to a real Excel number where possible.

    Blank -> None (an empty cell, not the literal string ""). Otherwise try
    int, then float, else keep the original string (ids like `gh/12345`,
    booleans like `True`, free text).
    """
    if raw == "":
        return None
    try:
        return int(raw)
    except ValueError:
        pass
    try:
        return float(raw)
    except ValueError:
        return raw


def _write_sheet(ws: Worksheet, csv_path: Path) -> int:
    """Write one CSV's contents into `ws`, styling row 1 as the header.
    Returns the number of data rows written (0 if the CSV is missing/empty).
    """
    if not csv_path.exists():
        return 0
    with open(csv_path, encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        return 0

    header, *data_rows = rows
    ws.append(header)
    for row in data_rows:
        ws.append([_cell_value(v) for v in row])

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return len(data_rows)


def _repo_url(repo: str, repo_id: str) -> str | None:
    """The repo's home page, derived from its platform-qualified id:
    `gh/<n>` → github.com, bare `gl/<n>` → gitlab.com, `gl/<nickname>-<n>` →
    that self-hosted GitLab instance (nickname resolved via
    gitlab_client.HOST_NICKNAMES). Unknown/blank id → no link."""
    if repo_id.startswith("gh/"):
        return f"https://github.com/{repo}"
    if repo_id.startswith("gl/"):
        from src.sources.gitlab.gitlab_client import HOST_NICKNAMES
        rest = repo_id[3:]
        if rest.isdigit():
            return f"https://gitlab.com/{repo}"
        nickname = rest.rsplit("-", 1)[0]
        host = {n: h for h, n in HOST_NICKNAMES.items() if n}.get(nickname)
        return f"https://{host}/{repo}" if host else None
    return None


def _decorate_repos_sheet(ws: Worksheet) -> None:
    """repos-only dressing (the reviewed design): hyperlink each `repo` cell
    to the repo's home page (host from `repo_id`), apply the static fills,
    center/bold the metric cells, add the conditional formats per column
    group, and set the fixed column widths."""
    headers = {c.value: c.column for c in ws[1]}  # name -> 1-based col index
    repo_col, id_col = headers.get("repo"), headers.get("repo_id")
    fill_cols = {headers[n]: f for n, f in REPOS_STATIC_FILLS.items() if n in headers}
    center_cols = {headers[n] for n in REPOS_CENTERED_COLS if n in headers}
    bold_cols = {headers[n] for n in REPOS_BOLD_COLS if n in headers}
    for row in ws.iter_rows(min_row=2):
        if repo_col and id_col:
            cell = row[repo_col - 1]
            url = _repo_url(str(cell.value or ""), str(row[id_col - 1].value or ""))
            if url:
                cell.hyperlink = url
                cell.style = "Hyperlink"
        for col, fill in fill_cols.items():
            row[col - 1].fill = fill
        for col in center_cols:
            row[col - 1].alignment = CENTER
        for col in bold_cols:
            row[col - 1].font = BOLD
    _add_repos_conditional_formats(ws, headers)
    for name, width in REPOS_COLUMN_WIDTHS.items():
        if name in headers:
            ws.column_dimensions[get_column_letter(headers[name])].width = width


def _add_repos_conditional_formats(ws: Worksheet, headers: dict) -> None:
    """Conditional formats on the repos sheet, one per column group, each
    over the column's data range (row 2..max_row):

      - value columns: 3-color scale, 0=red → 50=yellow → 100=green;
      - risk columns: the same scale reversed — higher risk reads red;
      - `score`: 2-color scale, 0=white → 100=blue;
      - boolean columns: "True" in dark-green text, "False" in dark-red
        (font only, so `eligible`'s static purple fill shows through);
      - `priority`: light-blue fill on non-empty cells.
    """
    def data_range(name: str) -> tuple[str, str] | tuple[None, None]:
        col = headers.get(name)
        if not col or ws.max_row < 2:
            return None, None
        letter = get_column_letter(col)
        return letter, f"{letter}2:{letter}{ws.max_row}"

    for name in REPOS_VALUE_SCALE_COLS:
        _, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color=SCALE_RED,
                mid_type="num", mid_value=50, mid_color=SCALE_YELLOW,
                end_type="num", end_value=100, end_color=SCALE_GREEN))
    for name in REPOS_RISK_SCALE_COLS:
        _, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="num", start_value=0, start_color=SCALE_GREEN,
                mid_type="num", mid_value=50, mid_color=SCALE_YELLOW,
                end_type="num", end_value=100, end_color=SCALE_RED))
    _, rng = data_range(REPOS_SCORE_COL)
    if rng:
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            end_type="num", end_value=100, end_color=SCORE_BLUE))
    for name in REPOS_BOOL_COLS:
        _, rng = data_range(name)
        if rng:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"True"'], font=BOOL_TRUE_FONT))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"False"'], font=BOOL_FALSE_FONT))
    letter, rng = data_range(REPOS_PRIORITY_COL)
    if rng:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"NOT(ISBLANK({letter}2))"], fill=PRIORITY_FILL))


def _md_cell(raw: str) -> tuple[int | float | str | None, str | None]:
    """One markdown table cell → (Excel value, number format).

    Strips the markdown dressing (`**bold**`, `` `code` `` — including code
    spans inside prose), then coerces to REAL typed cells so Excel never
    flags "number stored as text": thousands-separated counts become ints
    (format `#,##0`), decimals become floats (`0.00`), and percent strings
    ("97.8%") become fractions with a percent format (`0.0%`, or `0%` when
    the source had no decimals). Prose stays text (no format); a
    blank/placeholder cell becomes an empty cell."""
    s = raw.strip().strip("*").replace("`", "").strip()
    if s in ("", "—", "·"):
        return None, None
    if s.endswith("%"):
        num = s[:-1].replace(",", "").strip()
        try:
            return float(num) / 100.0, ("0.0%" if "." in num else "0%")
        except ValueError:
            return s, None
    plain = s.replace(",", "")
    try:
        return int(plain), "#,##0"
    except ValueError:
        pass
    try:
        return float(plain), "0.00"
    except ValueError:
        return s, None


def _is_md_separator(line: str) -> bool:
    """A markdown table's |---|--:|---| alignment row (no letters/digits)."""
    body = line.strip().strip("|")
    return bool(body) and set(body) <= set("-:| ")


def _stats_markdown() -> str:
    """The stats tables as markdown, straight from the generator.

    `scripts/stats.py` computes every count from the live CSVs; its
    `markdown()` renderer is the one source of truth for the numbers on
    this sheet (the preview stats sheet no longer exists)."""
    import importlib.util
    spec = importlib.util.spec_from_file_location("_stats_gen", STATS_SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.markdown(mod.value_stats(), mod.risk_stats(), mod.eligibility_stats())


THIN = Side(style="thin", color="000000")


def _with_sides(cell, **sides) -> None:
    """Merge border sides onto a cell without clobbering the ones it has."""
    b = cell.border
    cell.border = Border(left=sides.get("left", b.left),
                         right=sides.get("right", b.right),
                         top=sides.get("top", b.top),
                         bottom=sides.get("bottom", b.bottom))


def _render_table(ws: Worksheet, block: list[list[str]],
                  top_row: int, start_col: int) -> int:
    """Render one parsed markdown table at (top_row, start_col); returns its
    last row. Handles typed cells + number formats, whole-row bold from
    `**` markers, `^` separator lines, numeric-header right-alignment and
    the thin box outline."""
    n_cols = max(len(cells) for cells in block)
    numeric_col = [True] * n_cols
    has_data_col = [False] * n_cols
    separator_rows: list[int] = []
    for j, cells in enumerate(block):
        row_idx = top_row + j
        first = cells[0].strip()
        if first.startswith("^"):
            separator_rows.append(row_idx)
            cells = [first.lstrip("^")] + cells[1:]
        bold = cells[0].strip().startswith("**")
        for k, (v, fmt) in enumerate(_md_cell(x) for x in cells):
            cell = ws.cell(row=row_idx, column=start_col + k, value=v)
            if fmt:
                cell.number_format = fmt
            if j == 0:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                if bold:
                    cell.font = Font(bold=True)
                if v is not None:
                    has_data_col[k] = True
                    if not isinstance(v, (int, float)):
                        numeric_col[k] = False
    for k in range(n_cols):
        if has_data_col[k] and numeric_col[k]:
            ws.cell(row=top_row, column=start_col + k).alignment = \
                Alignment(horizontal="right")
    # thin box outline + separator lines above marked rows
    last_row = top_row + len(block) - 1
    last_col = start_col + n_cols - 1
    for row in range(top_row, last_row + 1):
        _with_sides(ws.cell(row=row, column=start_col), left=THIN)
        _with_sides(ws.cell(row=row, column=last_col), right=THIN)
    for col in range(start_col, last_col + 1):
        _with_sides(ws.cell(row=top_row, column=col), top=THIN)
        _with_sides(ws.cell(row=last_row, column=col), bottom=THIN)
        for row in separator_rows:
            _with_sides(ws.cell(row=row, column=col), top=THIN)
    return last_row


BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)

# Stage banner fills (matched to the reviewed sheet): value blue, risk
# orange, eligibility purple; anything else falls back to the header navy.
BANNER_FILLS = {
    "Stage 1": _fill("5B9BD5"),
    "Stage 2": _fill("ED7D31"),
    "Stage 3": _fill("8064A2"),
}
BANNER_SPAN_COLS = 13  # fill A..M so the banner reads as a full-width bar


def _write_stats_sheet(ws: Worksheet, md_text: str) -> int:
    """Render the generator's markdown as the stats sheet.

    Layout rules (mirroring the reviewed design):
      - gridlines off; column A is an empty gutter, content starts at B;
      - `## …` headings are STAGE BANNERS — bold white on the header fill;
      - `### …` headings are bold sub-titles, printed once above their table;
      - a `<!-- beside -->` directive renders the NEXT table to the right of
        the previous one (same rows, one gap column) — used for the CPP
        ecosystem breakdown next to the value funnel;
      - `^` first-cell markers draw separator lines, `**` rows are bold.
    """
    ws.sheet_view.showGridLines = False
    lines = md_text.splitlines()
    heading = ""
    heading_level = 3
    heading_emitted = True
    beside = False
    prev = None  # (top_row, start_col, n_cols, last_row) of the previous table
    cursor = 2   # next free row (row 1 stays blank, like the column-A gutter)
    tables = 0
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith("#"):
            heading_level = len(line) - len(line.lstrip("#"))
            heading = line.lstrip("#").strip()
            if heading_level == 2:
                # Stage banner — standalone, emitted immediately, full-width.
                fill = next((f for k, f in BANNER_FILLS.items()
                             if heading.startswith(k)), HEADER_FILL)
                for col in range(1, BANNER_SPAN_COLS + 1):
                    ws.cell(row=cursor, column=col).fill = fill
                cell = ws.cell(row=cursor, column=2, value=heading)
                cell.font = BANNER_FONT
                cursor += 3  # banner + two blank rows
                heading_emitted = True
            else:
                heading_emitted = False
            i += 1
            continue
        if line == "<!-- beside -->":
            beside = True
            i += 1
            continue
        if not line.startswith("|"):
            i += 1
            continue
        block = []
        while i < len(lines) and lines[i].strip().startswith("|"):
            if not _is_md_separator(lines[i]):
                block.append([x for x in lines[i].strip().strip("|").split("|")])
            i += 1
        if not block:
            continue

        if beside and prev is not None:
            top_row, start_col = prev[0], prev[1] + prev[2] + 1
            last = _render_table(ws, block, top_row, start_col)
            cursor = max(cursor, last + 2)
            prev = (top_row, start_col, max(len(r) for r in block), last)
            beside = False
            tables += 1
            continue

        if not heading_emitted:
            cell = ws.cell(row=cursor, column=2, value=heading)
            cell.font = SECTION_FONT
            cursor += 1  # sub-title sits directly above its table
            heading_emitted = True
        top_row = cursor
        last = _render_table(ws, block, top_row, 2)
        prev = (top_row, 2, max(len(r) for r in block), last)
        cursor = last + 3  # two blank rows between blocks
        tables += 1

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 60  # metric labels carry the folded descriptions
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 26  # beside-table label column
    for col in "JKLM":
        ws.column_dimensions[col].width = 12
    return tables


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet; we name our own

    for name, path in SHEETS:
        ws = wb.create_sheet(name)
        n = _write_sheet(ws, path)
        if name == "repos" and n:
            _decorate_repos_sheet(ws)
        console.print(f"  [cyan]{name}[/cyan]: {n:,} rows ← {path}")

    ws = wb.create_sheet("stats")
    n = _write_stats_sheet(ws, _stats_markdown())
    console.print(f"  [cyan]stats[/cyan]: {n} tables ← scripts/stats.py (live CSVs)")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)


def main() -> None:
    console.print("[bold]Building preview.xlsx...[/bold]\n")
    build()
    console.print(f"\n[dim]Wrote {OUTPUT_FILE}[/dim]")


if __name__ == "__main__":
    main()
